"""
xi_calibration_test.py  -  xi Floor Calibration Test Workbook

Tests 6 xi floor configurations to find a sweet spot where 3-5
developing countries enter the efficiency-adjusted top 15.

Data source: flop_trade_model_v26.docx Table A1 (index 10) & A2 (index 11)
Output: xi_calibration_test.xlsx (4 sheets: Rankings, Top20, Summary, DevDetail)

Does NOT modify the docx.
"""

import os
import shutil
import tempfile
import pathlib

from docx import Document
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =====================================================================
# PATHS
# =====================================================================
DOCX_SRC = pathlib.Path(
    r"F:\onedrive\__documents\papers\FLOPsExport\Documents"
    r"\flop_trade_model_v26.docx"
)
DATA_DIR = pathlib.Path(r"F:\onedrive\__documents\papers\FLOPsExport\Data")
OUT_XLSX = DATA_DIR / "xi_calibration_test.xlsx"

# =====================================================================
# OECD SET (38 members, using Table A1 spellings)
# Protocol lists 37; Croatia (acceded Dec 2024) added here.
# =====================================================================
OECD = {
    'Australia', 'Austria', 'Belgium', 'Canada', 'Chile', 'Colombia',
    'Croatia', 'Czechia', 'Denmark', 'Estonia', 'Finland', 'France',
    'Germany', 'Greece', 'Hungary', 'Iceland', 'Ireland', 'Israel',
    'Italy', 'Japan', 'Latvia', 'Lithuania', 'Luxembourg', 'Mexico',
    'Netherlands', 'New Zealand', 'Norway', 'Poland', 'Portugal',
    'Slovakia', 'Slovenia', 'South Korea', 'Spain', 'Sweden',
    'Switzerland', 'Turkey', 'United Kingdom', 'United States of Am.',
}

# Name normalization (docx variants -> canonical)
NAME_MAP = {
    'United States of A.': 'United States of Am.',
    'UK': 'United Kingdom',
    'UAE': 'United Arab Emirates',
    'N. Macedonia': 'North Macedonia',
}

# Tracked developing countries for Summary sheet
TRACKED_DEV = [
    'Kyrgyzstan', 'Kazakhstan', 'Georgia', 'Ethiopia', 'Uzbekistan',
    'Egypt', 'China', 'Vietnam', 'Montenegro', 'Kosovo',
]

# =====================================================================
# CONFIGURATIONS
# =====================================================================
CONFIGS = [
    {'label': 'BL', 'name': 'Config 0 (BL)', 'floor': 0.50,
     'type': 'linear'},
    {'label': 'C1', 'name': 'Config 1',      'floor': 0.85,
     'type': 'linear'},
    {'label': 'C2', 'name': 'Config 2',      'floor': 0.90,
     'type': 'linear'},
    {'label': 'C3', 'name': 'Config 3',      'floor': 0.925,
     'type': 'linear'},
    {'label': 'C4', 'name': 'Config 4',      'floor': 0.95,
     'type': 'linear'},
    {'label': 'C5', 'name': 'Config 5 (C)',  'exponent': 0.10,
     'type': 'concave'},
]


# =====================================================================
# HELPERS
# =====================================================================
def spearman_rho(ranks_a, ranks_b):
    """Spearman rank correlation coefficient."""
    n = len(ranks_a)
    if n < 2:
        return 0.0
    d_sq = sum((a - b) ** 2 for a, b in zip(ranks_a, ranks_b))
    return 1 - 6 * d_sq / (n * (n ** 2 - 1))


def normalize_name(name):
    """Apply name map and strip whitespace."""
    name = name.strip()
    return NAME_MAP.get(name, name)


# =====================================================================
# STEP 1: EXTRACT DATA FROM DOCX
# =====================================================================
def extract_from_docx():
    """
    Extract xi_eff from Table A1 (index 10, col 7)
    and cr_cost from Table A2 (index 11, col 4).
    Back-compute xi_raw = 2 * xi_eff - 1.
    Returns list of country dicts.
    """
    # Copy to temp to avoid OneDrive/Word lock
    tmp_dir = tempfile.mkdtemp(prefix="xi_cal_")
    tmp_docx = os.path.join(tmp_dir, "flop_v26.docx")
    shutil.copy2(str(DOCX_SRC), tmp_docx)

    try:
        doc = Document(tmp_docx)
        tables = doc.tables

        # --- Table A1 (index 10): 1 header + 85 data rows ---
        t_a1 = tables[10]
        a1_data = {}  # name -> xi_eff
        for r in range(1, len(t_a1.rows)):
            name_raw = t_a1.cell(r, 0).text.strip()
            xi_str = t_a1.cell(r, 7).text.strip()
            if not name_raw or not xi_str:
                continue
            name = normalize_name(name_raw)
            a1_data[name] = float(xi_str)

        print(f"  Table A1: extracted {len(a1_data)} countries")

        # --- Table A2 (index 11): 2 header rows + 85 data rows ---
        t_a2 = tables[11]
        a2_data = {}  # name -> cr_cost
        for r in range(2, len(t_a2.rows)):
            name_raw = t_a2.cell(r, 0).text.strip()
            cr_str = (t_a2.cell(r, 4).text.strip()
                      .replace('$', '').replace(',', ''))
            if not name_raw or not cr_str:
                continue
            name = normalize_name(name_raw)
            a2_data[name] = float(cr_str)

        print(f"  Table A2: extracted {len(a2_data)} countries")

    finally:
        os.unlink(tmp_docx)
        os.rmdir(tmp_dir)

    # Match countries between tables
    common = set(a1_data.keys()) & set(a2_data.keys())
    only_a1 = set(a1_data.keys()) - common
    only_a2 = set(a2_data.keys()) - common

    # First-10-char fallback for unmatched names
    matched_fallback = {}
    for n1 in list(only_a1):
        for n2 in list(only_a2):
            if n1[:10] == n2[:10]:
                matched_fallback[n1] = n2
                break

    if matched_fallback:
        print(f"  Fallback matches: {matched_fallback}")
        for n1, n2 in matched_fallback.items():
            a2_data[n1] = a2_data.pop(n2)
            common.add(n1)
            only_a1.discard(n1)
            only_a2.discard(n2)

    if only_a1 or only_a2:
        print(f"  WARNING: Unmatched in A1: {only_a1}")
        print(f"  WARNING: Unmatched in A2: {only_a2}")

    # Build country records (preserve Table A1 order for stable tie-breaking)
    a1_ordered = []
    for r in range(1, tables[10].rows.__len__()):
        name_raw = tables[10].cell(r, 0).text.strip()
        if name_raw:
            a1_ordered.append(normalize_name(name_raw))

    countries = []
    for name in a1_ordered:
        if name not in common:
            continue
        xi_eff = a1_data[name]
        cr_cost = a2_data[name]
        # Back-compute: xi_eff = 0.50 + 0.50 * xi_raw  =>  xi_raw = 2*xi_eff - 1
        xi_raw = 2.0 * xi_eff - 1.0
        xi_raw = max(0.0, min(1.0, xi_raw))
        is_dev = name not in OECD
        countries.append({
            'name': name,
            'xi_eff_orig': xi_eff,
            'cr_cost': cr_cost,
            'xi_raw': xi_raw,
            'is_dev': is_dev,
        })

    n_oecd = sum(1 for c in countries if not c['is_dev'])
    n_dev = sum(1 for c in countries if c['is_dev'])
    print(f"  Matched: {len(countries)} countries "
          f"(OECD: {n_oecd}, Developing: {n_dev})")

    return countries


# =====================================================================
# STEP 2: COMPUTE CONFIGURATIONS
# =====================================================================
def compute_configs(countries):
    """
    For each configuration, compute xi_eff, adjusted_cost, and rank.
    """
    # cr_rank (rank by cr_cost ascending)
    sorted_cr = sorted(range(len(countries)),
                       key=lambda i: countries[i]['cr_cost'])
    for rank, idx in enumerate(sorted_cr, 1):
        countries[idx]['cr_rank'] = rank

    for cfg in CONFIGS:
        label = cfg['label']

        for c in countries:
            xi_raw = c['xi_raw']

            if cfg['type'] == 'linear':
                floor = cfg['floor']
                xi_eff = floor + (1.0 - floor) * xi_raw
            else:  # concave
                exp = cfg['exponent']
                xi_eff = xi_raw ** exp if xi_raw > 0 else 0.0

            adj_cost = c['cr_cost'] / xi_eff if xi_eff > 0 else 999.0
            c[f'xi_eff_{label}'] = xi_eff
            c[f'adj_cost_{label}'] = adj_cost

        # Rank by adj_cost ascending (stable sort preserves extraction order)
        sorted_idx = sorted(range(len(countries)),
                            key=lambda i: countries[i][f'adj_cost_{label}'])
        for rank, idx in enumerate(sorted_idx, 1):
            countries[idx][f'rank_{label}'] = rank


# =====================================================================
# STEP 3: VERIFICATION (10 mandatory checks)
# =====================================================================
def verify(countries):
    """Run 10 mandatory verification checks."""
    ok = True
    print(f"\n{'='*60}")
    print("VERIFICATION CHECKS")
    print(f"{'='*60}")

    # [1] 85 countries, no duplicates
    n = len(countries)
    names = [c['name'] for c in countries]
    n_unique = len(set(names))
    if n != 85 or n_unique != 85:
        print(f"  FAIL [1]: {n} countries, {n_unique} unique (expected 85)")
        ok = False
    else:
        print("  OK   [1]: 85 countries, no duplicates")

    # [2] All xi_eff <= 1.00
    for cfg in CONFIGS:
        label = cfg['label']
        vals = [c[f'xi_eff_{label}'] for c in countries]
        if any(v > 1.001 for v in vals):
            print(f"  FAIL [2-{label}]: xi_eff > 1.00")
            ok = False
        else:
            print(f"  OK   [2-{label}]: xi_eff in "
                  f"[{min(vals):.4f}, {max(vals):.4f}]")

    # [3] All xi_eff >= config floor
    for cfg in CONFIGS:
        label = cfg['label']
        vals = [c[f'xi_eff_{label}'] for c in countries]
        if cfg['type'] == 'linear':
            floor = cfg['floor']
            if any(v < floor - 0.001 for v in vals):
                print(f"  FAIL [3-{label}]: xi_eff < floor {floor}")
                ok = False
            else:
                print(f"  OK   [3-{label}]: all xi_eff >= floor {floor}")
        else:
            xi_raws = [c['xi_raw'] for c in countries]
            min_raw = min(xi_raws)
            min_expected = min_raw ** cfg['exponent'] if min_raw > 0 else 0
            print(f"  OK   [3-{label}]: concave, min xi_eff = "
                  f"{min(vals):.4f} (floor ~{min_expected:.4f})")

    # [4] HARD CHECK: Baseline top 5
    top5_bl = sorted(countries, key=lambda c: c['rank_BL'])[:5]
    expected_top5 = ['Canada', 'Norway', 'Finland', 'Sweden', 'South Africa']
    actual_top5 = [c['name'] for c in top5_bl]
    if actual_top5 == expected_top5:
        print(f"  OK   [4]: Baseline top 5 = {actual_top5}")
    else:
        print("  FAIL [4]: Baseline top 5 mismatch!")
        print(f"    Expected: {expected_top5}")
        print(f"    Actual:   {actual_top5}")
        print("    Top 10 details:")
        for c in sorted(countries, key=lambda c: c['rank_BL'])[:10]:
            print(f"      {c['rank_BL']:2d}. {c['name']:30s} "
                  f"cr={c['cr_cost']:.4f} xi_eff={c['xi_eff_BL']:.4f} "
                  f"adj={c['adj_cost_BL']:.4f}")
        ok = False

    # [5] Rankings 1-85, no gaps
    for cfg in CONFIGS:
        label = cfg['label']
        ranks = sorted(c[f'rank_{label}'] for c in countries)
        if ranks != list(range(1, 86)):
            print(f"  FAIL [5-{label}]: ranks not 1-85")
            ok = False
        else:
            print(f"  OK   [5-{label}]: ranks 1-85 complete")
    cr_ranks = sorted(c['cr_rank'] for c in countries)
    if cr_ranks != list(range(1, 86)):
        print("  FAIL [5-cr]: cr_rank not 1-85")
        ok = False
    else:
        print("  OK   [5-cr]: cr_rank 1-85 complete")

    # [6] cr_cost identical across configs (trivially true: single source)
    print("  OK   [6]: cr_cost identical across configs (single source)")

    # [7] xi_raw=1.00 countries -> xi_eff=1.00 under all configs
    xi1 = [c for c in countries if abs(c['xi_raw'] - 1.0) < 0.005]
    all_ok_7 = True
    for c in xi1:
        for cfg in CONFIGS:
            label = cfg['label']
            xi_eff = c[f'xi_eff_{label}']
            if abs(xi_eff - 1.0) > 0.005:
                print(f"  FAIL [7]: {c['name']} {label} "
                      f"xi_eff={xi_eff:.4f} (xi_raw={c['xi_raw']:.4f})")
                all_ok_7 = False
    if all_ok_7:
        names_1 = [c['name'] for c in xi1]
        print(f"  OK   [7]: {len(xi1)} countries with xi_raw=1.00 -> "
              f"xi_eff=1.00 under all configs")
        print(f"            {names_1}")
    else:
        ok = False

    # [8] Spearman rho (config vs cr_rank) increases with floor
    cr_ranks_list = [c['cr_rank'] for c in countries]
    rhos = []
    for cfg in CONFIGS:
        label = cfg['label']
        cfg_ranks = [c[f'rank_{label}'] for c in countries]
        rho = spearman_rho(cr_ranks_list, cfg_ranks)
        rhos.append((label, rho))
    rho_str = "  ".join(f"{lbl}={rho:.4f}" for lbl, rho in rhos)
    print(f"  Spearman rho (vs cr_rank): {rho_str}")

    linear_rhos = [r for lbl, r in rhos if lbl != 'C5']
    if all(a <= b + 0.001 for a, b in zip(linear_rhos, linear_rhos[1:])):
        print("  OK   [8]: Spearman rho increases with floor (linear)")
    else:
        print("  WARN [8]: Spearman rho NOT strictly increasing")

    # [9] Spearman(BL vs BL) = 1.0
    bl_ranks = [c['rank_BL'] for c in countries]
    rho_self = spearman_rho(bl_ranks, bl_ranks)
    if abs(rho_self - 1.0) < 0.0001:
        print(f"  OK   [9]: Spearman(BL vs BL) = {rho_self:.4f}")
    else:
        print(f"  FAIL [9]: Spearman(BL vs BL) = {rho_self:.4f}")
        ok = False

    # [10] Dev in top 15 weakly increases with floor
    dev_counts = []
    for cfg in CONFIGS:
        label = cfg['label']
        n_dev = sum(1 for c in countries
                    if c[f'rank_{label}'] <= 15 and c['is_dev'])
        dev_counts.append((label, n_dev))
    dev_str = "  ".join(f"{lbl}={n}" for lbl, n in dev_counts)
    print(f"  Dev in top 15: {dev_str}")

    linear_devs = [n for lbl, n in dev_counts if lbl != 'C5']
    if all(a <= b for a, b in zip(linear_devs, linear_devs[1:])):
        print("  OK   [10]: Dev count weakly increases with floor")
    else:
        print("  WARN [10]: Dev count NOT weakly increasing")

    print(f"{'='*60}\n")
    return ok


# =====================================================================
# STEP 4: WRITE EXCEL (4 sheets)
# =====================================================================
COST_FMT = '$#,##0.00'
ADJ_COST_FMT = '$#,##0.0000'
XI_FMT = '0.0000'
RANK_FMT = '0'
SPEARMAN_FMT = '0.0000'

HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3",
                          fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                         fill_type="solid")
BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE",
                        fill_type="solid")
BOLD = Font(bold=True)
BOLD_10 = Font(bold=True, size=10)


def _header(ws, headers, widths=None, row=1):
    """Write header row with formatting."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def _cell(ws, row, col, value, fmt=None, font=None, fill=None):
    """Write and format a cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if fmt:
        cell.number_format = fmt
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    return cell


def write_excel(countries):
    """Write xi_calibration_test.xlsx with 4 sheets."""
    wb = openpyxl.Workbook()

    # ================================================================
    # Sheet 1: Rankings (85 rows sorted by cr_cost)
    # ================================================================
    ws = wb.active
    ws.title = "Rankings"

    headers = ["Country", "cr_cost", "cr_rank", "xi_raw"]
    widths = [28, 11, 8, 10]
    for cfg in CONFIGS:
        lbl = cfg['label']
        headers += [f"xi_eff ({lbl})", f"adj_cost ({lbl})", f"rank ({lbl})"]
        widths += [12, 14, 10]
    _header(ws, headers, widths)

    sorted_data = sorted(countries, key=lambda c: c['cr_cost'])
    for i, c in enumerate(sorted_data, 2):
        fill = GREEN_FILL if c['is_dev'] else None
        col = 1
        _cell(ws, i, col, c['name'], fill=fill)
        col += 1
        _cell(ws, i, col, c['cr_cost'], COST_FMT, fill=fill)
        col += 1
        _cell(ws, i, col, c['cr_rank'], RANK_FMT, fill=fill)
        col += 1
        _cell(ws, i, col, c['xi_raw'], XI_FMT, fill=fill)
        col += 1
        for cfg in CONFIGS:
            lbl = cfg['label']
            _cell(ws, i, col, c[f'xi_eff_{lbl}'], XI_FMT, fill=fill)
            col += 1
            _cell(ws, i, col, c[f'adj_cost_{lbl}'], ADJ_COST_FMT, fill=fill)
            col += 1
            _cell(ws, i, col, c[f'rank_{lbl}'], RANK_FMT, fill=fill)
            col += 1

    # ================================================================
    # Sheet 2: Top20 (side-by-side, top 20 per config)
    # ================================================================
    ws = wb.create_sheet("Top20")

    headers = []
    widths = []
    for cfg in CONFIGS:
        lbl = cfg['label']
        headers += [f"Rank ({lbl})", f"Country ({lbl})",
                    f"adj_cost ({lbl})"]
        widths += [10, 28, 14]
    _header(ws, headers, widths)

    for cfg_idx, cfg in enumerate(CONFIGS):
        lbl = cfg['label']
        top20 = sorted(countries,
                       key=lambda c: c[f'rank_{lbl}'])[:20]
        base_col = cfg_idx * 3 + 1
        for i, c in enumerate(top20, 2):
            star = " \u2605" if c['is_dev'] else ""
            _cell(ws, i, base_col, c[f'rank_{lbl}'], RANK_FMT)
            _cell(ws, i, base_col + 1, c['name'] + star)
            _cell(ws, i, base_col + 2,
                  c[f'adj_cost_{lbl}'], ADJ_COST_FMT)

    # ================================================================
    # Sheet 3: Summary
    # ================================================================
    ws = wb.create_sheet("Summary")

    # Determine sweet-spot columns (3-8 dev in top 15)
    dev_top15 = {}
    for cfg in CONFIGS:
        lbl = cfg['label']
        n = sum(1 for c in countries
                if c[f'rank_{lbl}'] <= 15 and c['is_dev'])
        dev_top15[lbl] = n
    sweet_cols = {lbl for lbl, n in dev_top15.items() if 3 <= n <= 8}

    # Build metric rows
    metric_rows = []

    # xi_eff range
    row = ['xi_eff range']
    for cfg in CONFIGS:
        lbl = cfg['label']
        vals = [c[f'xi_eff_{lbl}'] for c in countries]
        row.append(f"[{min(vals):.4f}, {max(vals):.4f}]")
    metric_rows.append(row)

    # xi_eff min
    row = ['xi_eff min']
    for cfg in CONFIGS:
        lbl = cfg['label']
        vals = [c[f'xi_eff_{lbl}'] for c in countries]
        row.append(min(vals))
    metric_rows.append(row)

    # Max penalty (%)
    row = ['Max penalty (%)']
    for cfg in CONFIGS:
        lbl = cfg['label']
        vals = [c[f'xi_eff_{lbl}'] for c in countries]
        penalty = (1.0 - min(vals)) * 100
        row.append(f"{penalty:.1f}%")
    metric_rows.append(row)

    # Dev in top 10
    row = ['Dev in top 10']
    for cfg in CONFIGS:
        lbl = cfg['label']
        n = sum(1 for c in countries
                if c[f'rank_{lbl}'] <= 10 and c['is_dev'])
        row.append(n)
    metric_rows.append(row)

    # Dev in top 15
    row = ['Dev in top 15']
    for cfg in CONFIGS:
        row.append(dev_top15[cfg['label']])
    metric_rows.append(row)

    # Dev in top 20
    row = ['Dev in top 20']
    for cfg in CONFIGS:
        lbl = cfg['label']
        n = sum(1 for c in countries
                if c[f'rank_{lbl}'] <= 20 and c['is_dev'])
        row.append(n)
    metric_rows.append(row)

    # Spearman rho (vs cr_rank)
    cr_ranks = [c['cr_rank'] for c in countries]
    row = ['Spearman rho (vs cr_rank)']
    for cfg in CONFIGS:
        lbl = cfg['label']
        cfg_ranks = [c[f'rank_{lbl}'] for c in countries]
        row.append(spearman_rho(cr_ranks, cfg_ranks))
    metric_rows.append(row)

    # Spearman rho (vs BL)
    bl_ranks = [c['rank_BL'] for c in countries]
    row = ['Spearman rho (vs BL)']
    for cfg in CONFIGS:
        lbl = cfg['label']
        cfg_ranks = [c[f'rank_{lbl}'] for c in countries]
        row.append(spearman_rho(bl_ranks, cfg_ranks))
    metric_rows.append(row)

    # Write metrics header + data
    met_headers = ['Metric'] + [cfg['name'] for cfg in CONFIGS]
    met_widths = [26] + [18] * len(CONFIGS)
    _header(ws, met_headers, met_widths)

    for r, row_data in enumerate(metric_rows, 2):
        for c_idx, val in enumerate(row_data):
            cfg_lbl = CONFIGS[c_idx - 1]['label'] if c_idx > 0 else None
            fill = BLUE_FILL if cfg_lbl in sweet_cols else None
            if isinstance(val, float):
                fmt = (SPEARMAN_FMT
                       if 'Spearman' in str(row_data[0]) else XI_FMT)
                _cell(ws, r, c_idx + 1, val, fmt=fmt, fill=fill)
            else:
                _cell(ws, r, c_idx + 1, val, fill=fill)

    # Tracked country ranks
    track_start = 2 + len(metric_rows) + 2
    _cell(ws, track_start, 1, "Tracked Country Ranks", font=BOLD_10)
    track_headers = ['Country'] + [cfg['name'] for cfg in CONFIGS]
    for c_idx, h in enumerate(track_headers):
        cell = ws.cell(row=track_start + 1, column=c_idx + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for t_idx, tname in enumerate(TRACKED_DEV):
        row_num = track_start + 2 + t_idx
        c = next((x for x in countries if x['name'] == tname), None)
        if c is None:
            _cell(ws, row_num, 1, f"{tname} (not found)")
            continue
        _cell(ws, row_num, 1, tname)
        for cfg_idx, cfg in enumerate(CONFIGS):
            lbl = cfg['label']
            rank = c[f'rank_{lbl}']
            fill = GREEN_FILL if rank <= 15 else None
            _cell(ws, row_num, cfg_idx + 2, rank, RANK_FMT, fill=fill)

    # ================================================================
    # Sheet 4: DevDetail (developing countries)
    # ================================================================
    ws = wb.create_sheet("DevDetail")

    dev_countries = [c for c in countries if c['is_dev']]
    # Sort by Config 2 (C2) rank
    dev_countries.sort(key=lambda c: c['rank_C2'])

    headers = ["Country", "cr_cost", "cr_rank", "xi_raw"]
    widths = [28, 11, 8, 10]
    for cfg in CONFIGS:
        lbl = cfg['label']
        headers.append(f"rank ({lbl})")
        widths.append(10)
    headers += ["best_rank", "best_config"]
    widths += [10, 16]
    _header(ws, headers, widths)

    for i, c in enumerate(dev_countries, 2):
        # Find best rank across configs
        best_rank = 86
        best_cfg = ""
        for cfg in CONFIGS:
            lbl = cfg['label']
            r = c[f'rank_{lbl}']
            if r < best_rank:
                best_rank = r
                best_cfg = cfg['name']

        is_top15 = best_rank <= 15
        font = BOLD if is_top15 else None

        col = 1
        _cell(ws, i, col, c['name'], font=font)
        col += 1
        _cell(ws, i, col, c['cr_cost'], COST_FMT, font=font)
        col += 1
        _cell(ws, i, col, c['cr_rank'], RANK_FMT, font=font)
        col += 1
        _cell(ws, i, col, c['xi_raw'], XI_FMT, font=font)
        col += 1
        for cfg in CONFIGS:
            lbl = cfg['label']
            r = c[f'rank_{lbl}']
            fill = GREEN_FILL if r <= 15 else None
            _cell(ws, i, col, r, RANK_FMT, font=font, fill=fill)
            col += 1
        _cell(ws, i, col, best_rank, RANK_FMT, font=font)
        col += 1
        _cell(ws, i, col, best_cfg, font=font)

    # Save
    wb.save(str(OUT_XLSX))
    print(f"  Saved: {OUT_XLSX}")


# =====================================================================
# MAIN
# =====================================================================
def main():
    print("=" * 60)
    print("xi_calibration_test.py  -  xi Floor Calibration Test")
    print("=" * 60)

    # Step 1: Extract from docx
    print("\n[1] Extracting data from docx...")
    countries = extract_from_docx()

    # Quick data summary
    xi_raws = [c['xi_raw'] for c in countries]
    print(f"\n  xi_raw range: [{min(xi_raws):.4f}, {max(xi_raws):.4f}]")
    min_c = min(countries, key=lambda c: c['xi_raw'])
    print(f"  Min xi_raw: {min_c['name']} = {min_c['xi_raw']:.4f}")
    n_xi1 = sum(1 for c in countries if abs(c['xi_raw'] - 1.0) < 0.005)
    print(f"  Countries with xi_raw ~= 1.00: {n_xi1}")

    cr_costs = [c['cr_cost'] for c in countries]
    print(f"  cr_cost range: "
          f"[${min(cr_costs):.2f}, ${max(cr_costs):.2f}]")

    # Step 2: Compute configurations
    print("\n[2] Computing 6 configurations...")
    compute_configs(countries)

    # Step 3: Verification
    ok = verify(countries)
    if not ok:
        print("*** VERIFICATION FAILED — see details above ***")
        print("    Writing Excel anyway for debugging.\n")

    # Step 4: Write Excel
    print("[4] Writing Excel...")
    write_excel(countries)

    print("\nDone.")


if __name__ == "__main__":
    main()
