"""
form_b_simulations.py  -  Form B Simulations with (omega, floor) Grid

Tests 20 configurations of the local-cost adjustment formula:
  Form B:  c_adj = rho + (c_cr - rho) / xi_eff
  Form A:  c_adj = c_cr / xi_eff  (reference only)

Data sources:
  - cr_cost:    xi_calibration_test.xlsx (Rankings, col B)
  - G, R:       xi_scenarios.xlsx (Data, cols E, G)
  - xi_raw_v26: xi_scenarios.xlsx (Data, col M)

Output: form_b_simulations.xlsx (5 sheets)
"""

import math
import pathlib

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =====================================================================
# PATHS
# =====================================================================
DATA_DIR = pathlib.Path(r"F:\onedrive\__documents\papers\FLOPsExport\Data")
CAL_XLSX = DATA_DIR / "xi_calibration_test.xlsx"
SCEN_XLSX = DATA_DIR / "xi_scenarios.xlsx"
OUT_XLSX = DATA_DIR / "form_b_simulations.xlsx"

# =====================================================================
# CONSTANTS
# =====================================================================
RHO_DEFAULT = 1.36  # hardware cost $/hr

OECD = {
    'Australia', 'Austria', 'Belgium', 'Canada', 'Chile', 'Colombia',
    'Croatia', 'Czechia', 'Denmark', 'Estonia', 'Finland', 'France',
    'Germany', 'Greece', 'Hungary', 'Iceland', 'Ireland', 'Israel',
    'Italy', 'Japan', 'Latvia', 'Lithuania', 'Luxembourg', 'Mexico',
    'Netherlands', 'New Zealand', 'Norway', 'Poland', 'Portugal',
    'Slovakia', 'Slovenia', 'South Korea', 'Spain', 'Sweden',
    'Switzerland', 'Turkey', 'United Kingdom', 'United States of Am.',
}

# Name map: xi_scenarios.xlsx name -> xi_calibration_test.xlsx name
NAME_MAP_SCEN_TO_CAL = {
    'United States of America': 'United States of Am.',
}

TRACKED_DEV = ['Kyrgyzstan', 'China', 'Ethiopia', 'Montenegro', 'Kosovo']

# =====================================================================
# CONFIGURATIONS (20 total)
# =====================================================================
CONFIGS = []

# Main grid: 4 omega values x 3 floor values = 12 configs
for letter, omega in [('A', 0.85), ('B', 0.70), ('C', 0.50), ('D', 0.30)]:
    for idx, floor in [(1, 0.00), (2, 0.30), (3, 0.50)]:
        CONFIGS.append({
            'label': f'{letter}{idx}',
            'omega': omega,
            'floor': floor,
            'rho': RHO_DEFAULT,
            'form': 'B',
            'group': 'main',
        })

# Reference configs
CONFIGS += [
    {'label': 'REF_A', 'omega': 0.85, 'floor': 0.50,
     'rho': RHO_DEFAULT, 'form': 'A', 'group': 'ref',
     'desc': 'Form A, floor=0.50, w=0.85 (current v26)'},
    {'label': 'REF_B', 'omega': 0.85, 'floor': 0.90,
     'rho': RHO_DEFAULT, 'form': 'A', 'group': 'ref',
     'desc': 'Form A, floor=0.90, w=0.85'},
    {'label': 'REF_CR', 'form': 'CR', 'group': 'ref',
     'desc': 'Pure cost-recovery (no governance)'},
    {'label': 'REF_Bv26', 'floor': 0.50,
     'rho': RHO_DEFAULT, 'form': 'B_v26', 'group': 'ref',
     'desc': 'Form B with v26 xi_raw, floor=0.50'},
]

# Hardware sensitivity: all at omega=0.50, floor=0.30
for idx, rho in [(1, 1.30), (2, 1.33), (3, 1.39), (4, 1.42)]:
    CONFIGS.append({
        'label': f'H{idx}',
        'omega': 0.50,
        'floor': 0.30,
        'rho': rho,
        'form': 'B',
        'group': 'hw',
    })

MAIN_LABELS = [c['label'] for c in CONFIGS if c['group'] == 'main']
TOP20_LABELS = ['C2', 'C3', 'A3', 'REF_A', 'REF_CR']


# =====================================================================
# HELPERS
# =====================================================================
def spearman_rho(ranks_a, ranks_b):
    n = len(ranks_a)
    if n < 2:
        return 0.0
    d_sq = sum((a - b) ** 2 for a, b in zip(ranks_a, ranks_b))
    return 1 - 6 * d_sq / (n * (n ** 2 - 1))


# =====================================================================
# STEP 1: LOAD DATA
# =====================================================================
def load_data():
    """Load cr_cost, G, R, xi_raw_v26 from the two xlsx files."""
    # --- cr_cost from xi_calibration_test.xlsx ---
    wb_cal = openpyxl.load_workbook(str(CAL_XLSX), data_only=True)
    ws_cal = wb_cal['Rankings']
    cal_data = {}  # name -> cr_cost
    for r in range(2, ws_cal.max_row + 1):
        name = ws_cal.cell(r, 1).value
        cr = ws_cal.cell(r, 2).value
        if name and cr is not None:
            cal_data[name] = float(cr)
    wb_cal.close()
    print(f"  xi_calibration_test.xlsx: {len(cal_data)} countries")

    # --- G, R, xi_raw_v26 from xi_scenarios.xlsx ---
    wb_scen = openpyxl.load_workbook(str(SCEN_XLSX), data_only=True)
    ws_scen = wb_scen['Data']
    scen_data = {}  # name -> {G, R, xi_raw_v26}
    for r in range(2, ws_scen.max_row + 1):
        name_raw = ws_scen.cell(r, 1).value
        if not name_raw:
            continue
        # Map name to calibration-test convention
        name = NAME_MAP_SCEN_TO_CAL.get(name_raw, name_raw)
        G = ws_scen.cell(r, 5).value   # G_RoL
        R = ws_scen.cell(r, 7).value   # R_grid
        xi_raw_v26 = ws_scen.cell(r, 13).value  # xi_raw_v26
        if G is not None and R is not None:
            scen_data[name] = {
                'G': float(G),
                'R': float(R),
                'xi_raw_v26': float(xi_raw_v26) if xi_raw_v26 else 0.5,
            }
    wb_scen.close()
    print(f"  xi_scenarios.xlsx: {len(scen_data)} countries")

    # Match
    common = set(cal_data.keys()) & set(scen_data.keys())
    only_cal = set(cal_data.keys()) - common
    only_scen = set(scen_data.keys()) - common
    if only_cal or only_scen:
        print(f"  WARNING: Only in calibration: {only_cal}")
        print(f"  WARNING: Only in scenarios: {only_scen}")

    # Build unified dataset (preserve cal_data order: sorted by cr_cost)
    countries = []
    for name in sorted(common, key=lambda n: cal_data[n]):
        cr_cost = cal_data[name]
        G = scen_data[name]['G']
        R = scen_data[name]['R']
        xi_raw_v26 = scen_data[name]['xi_raw_v26']
        is_oecd = name in OECD
        countries.append({
            'name': name,
            'cr_cost': cr_cost,
            'G': G,
            'R': R,
            'xi_raw_v26': xi_raw_v26,
            'is_oecd': is_oecd,
            'is_dev': not is_oecd,
        })

    n_oecd = sum(1 for c in countries if c['is_oecd'])
    n_dev = sum(1 for c in countries if c['is_dev'])
    print(f"  Matched: {len(countries)} (OECD: {n_oecd}, Dev: {n_dev})")
    return countries


# =====================================================================
# STEP 2: COMPUTE CONFIGURATIONS
# =====================================================================
def compute_configs(countries):
    """Compute xi_raw, xi_eff, c_adj, rank, markup for all configs."""
    n = len(countries)

    # cr_rank
    cr_sorted = sorted(range(n), key=lambda i: countries[i]['cr_cost'])
    for rank, idx in enumerate(cr_sorted, 1):
        countries[idx]['cr_rank'] = rank

    for cfg in CONFIGS:
        label = cfg['label']
        form = cfg['form']

        for c in countries:
            if form == 'CR':
                # No governance adjustment
                c[f'xi_raw_{label}'] = 1.0
                c[f'xi_eff_{label}'] = 1.0
                c[f'c_adj_{label}'] = c['cr_cost']
            elif form == 'B_v26':
                # Form B with v26 xi_raw
                xi_raw = c['xi_raw_v26']
                floor = cfg['floor']
                xi_eff = floor + (1.0 - floor) * xi_raw
                rho = cfg['rho']
                c_adj = rho + (c['cr_cost'] - rho) / xi_eff
                c[f'xi_raw_{label}'] = xi_raw
                c[f'xi_eff_{label}'] = xi_eff
                c[f'c_adj_{label}'] = c_adj
            else:
                # Standard: compute xi_raw from G, R, omega
                omega = cfg['omega']
                floor = cfg['floor']
                rho = cfg['rho']
                G = max(c['G'], 1e-6)
                R = max(c['R'], 1e-6)
                xi_raw = (G ** omega) * (R ** (1.0 - omega))
                xi_eff = (floor + (1.0 - floor) * xi_raw
                          if floor > 0 else xi_raw)
                xi_eff = max(xi_eff, 1e-6)

                if form == 'A':
                    c_adj = c['cr_cost'] / xi_eff
                else:  # form == 'B'
                    c_adj = rho + (c['cr_cost'] - rho) / xi_eff

                c[f'xi_raw_{label}'] = xi_raw
                c[f'xi_eff_{label}'] = xi_eff
                c[f'c_adj_{label}'] = c_adj

            # Markup
            cr = c['cr_cost']
            c_adj_val = c[f'c_adj_{label}']
            c[f'markup_{label}'] = ((c_adj_val / cr) - 1.0) * 100 if cr > 0 else 0

        # Rank by c_adj ascending (stable sort)
        sorted_idx = sorted(range(n),
                            key=lambda i: countries[i][f'c_adj_{label}'])
        for rank, idx in enumerate(sorted_idx, 1):
            countries[idx][f'rank_{label}'] = rank


# =====================================================================
# STEP 3: VALIDATION
# =====================================================================
def validate(countries):
    ok = True
    print(f"\n{'='*60}")
    print("VALIDATION CHECKS")
    print(f"{'='*60}")

    # [1] 85 countries
    n = len(countries)
    if n != 85:
        print(f"  FAIL [1]: {n} countries (expected 85)")
        ok = False
    else:
        print(f"  OK   [1]: 85 countries")

    # [2] No duplicates
    names = [c['name'] for c in countries]
    if len(set(names)) != n:
        print(f"  FAIL [2]: duplicate names found")
        ok = False
    else:
        print(f"  OK   [2]: 85 unique names")

    # [3] REF_CR ranking = cr_rank
    mismatches = [c for c in countries
                  if c['rank_REF_CR'] != c['cr_rank']]
    if mismatches:
        print(f"  FAIL [3]: REF_CR rank != cr_rank for "
              f"{len(mismatches)} countries")
        for m in mismatches[:5]:
            print(f"    {m['name']}: REF_CR={m['rank_REF_CR']} "
                  f"cr={m['cr_rank']}")
        ok = False
    else:
        print(f"  OK   [3]: REF_CR rank = cr_rank for all")

    # [4] Perfect-governance countries (G=1, R=1) have xi_eff=1, markup=0
    perfect = [c for c in countries
               if abs(c['G'] - 1.0) < 0.01 and abs(c['R'] - 1.0) < 0.01]
    all_ok_4 = True
    for c in perfect:
        for cfg in CONFIGS:
            label = cfg['label']
            xi_eff = c[f'xi_eff_{label}']
            markup = c[f'markup_{label}']
            if abs(xi_eff - 1.0) > 0.01 or abs(markup) > 0.5:
                print(f"  FAIL [4]: {c['name']} {label}: "
                      f"xi_eff={xi_eff:.4f}, markup={markup:.2f}%")
                all_ok_4 = False
    if all_ok_4:
        pnames = [c['name'] for c in perfect]
        print(f"  OK   [4]: Perfect-governance countries ({len(perfect)}): "
              f"xi_eff=1, markup=0% — {pnames}")
    else:
        ok = False

    # [5] Monotonicity: for fixed omega, increasing floor weakly increases
    #     dev in top 15
    all_ok_5 = True
    for letter, omega in [('A', 0.85), ('B', 0.70),
                          ('C', 0.50), ('D', 0.30)]:
        devs = []
        for idx in [1, 2, 3]:
            label = f'{letter}{idx}'
            nd = sum(1 for c in countries
                     if c[f'rank_{label}'] <= 15 and c['is_dev'])
            devs.append(nd)
        if not all(a <= b for a, b in zip(devs, devs[1:])):
            print(f"  WARN [5]: omega={omega}: dev top 15 = {devs} "
                  f"(not monotone)")
            all_ok_5 = False
        else:
            print(f"  OK   [5-{letter}]: omega={omega}: "
                  f"dev top 15 = {devs} (monotone)")

    # [6] REF_A: 2 developing in top 15
    ref_a_dev = sum(1 for c in countries
                    if c['rank_REF_A'] <= 15 and c['is_dev'])
    if ref_a_dev == 2:
        print(f"  OK   [6]: REF_A has {ref_a_dev} dev in top 15")
    else:
        print(f"  WARN [6]: REF_A has {ref_a_dev} dev in top 15 "
              f"(expected 2)")

    # [7] Formula spot-check: Kyrgyzstan under C2
    kyr = next((c for c in countries if c['name'] == 'Kyrgyzstan'), None)
    if kyr:
        G_k = kyr['G']
        R_k = kyr['R']
        cr_k = kyr['cr_cost']
        xi_raw_exp = (G_k ** 0.50) * (R_k ** 0.50)
        xi_eff_exp = 0.30 + 0.70 * xi_raw_exp
        c_adj_exp = 1.36 + (cr_k - 1.36) / xi_eff_exp
        c_adj_act = kyr['c_adj_C2']
        if abs(c_adj_exp - c_adj_act) < 0.0001:
            print(f"  OK   [7]: Kyrgyzstan C2 spot-check: "
                  f"c_adj={c_adj_act:.4f} (expected {c_adj_exp:.4f})")
        else:
            print(f"  FAIL [7]: Kyrgyzstan C2: actual={c_adj_act:.4f} "
                  f"expected={c_adj_exp:.4f}")
            ok = False
        print(f"           G={G_k:.4f}, R={R_k:.4f}, cr={cr_k:.4f}")
        print(f"           xi_raw={xi_raw_exp:.4f}, "
              f"xi_eff={xi_eff_exp:.4f}")
    else:
        print(f"  FAIL [7]: Kyrgyzstan not found")
        ok = False

    print(f"{'='*60}\n")
    return ok


# =====================================================================
# STEP 4: WRITE EXCEL
# =====================================================================
COST_FMT = '$#,##0.00'
ADJ_FMT = '$#,##0.0000'
XI_FMT = '0.0000'
PCT_FMT = '0.0%'
MARKUP_FMT = '0.00"%"'
RANK_FMT = '0'
RHO_FMT = '0.0000'

HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3",
                           fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                          fill_type="solid")
BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE",
                          fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                           fill_type="solid")
BOLD = Font(bold=True)
BOLD_10 = Font(bold=True, size=10)


def _hdr(ws, headers, widths=None, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def _c(ws, row, col, value, fmt=None, font=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fmt:
        cell.number_format = fmt
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    return cell


def write_excel(countries):
    wb = openpyxl.Workbook()

    # ================================================================
    # Sheet 1: Data
    # ================================================================
    ws = wb.active
    ws.title = "Data"
    headers = ["Country", "cr_cost", "G", "R", "OECD", "developing",
               "cr_rank"]
    widths = [28, 11, 10, 10, 8, 10, 8]
    _hdr(ws, headers, widths)

    sorted_cr = sorted(countries, key=lambda c: c['cr_rank'])
    for i, c in enumerate(sorted_cr, 2):
        _c(ws, i, 1, c['name'])
        _c(ws, i, 2, c['cr_cost'], COST_FMT)
        _c(ws, i, 3, c['G'], XI_FMT)
        _c(ws, i, 4, c['R'], XI_FMT)
        _c(ws, i, 5, c['is_oecd'])
        _c(ws, i, 6, c['is_dev'])
        _c(ws, i, 7, c['cr_rank'], RANK_FMT)

    # ================================================================
    # Sheet 2: Rankings (all 20 configs side by side)
    # ================================================================
    ws = wb.create_sheet("Rankings")
    headers = ["Country"]
    widths = [28]
    for cfg in CONFIGS:
        lbl = cfg['label']
        headers += [f"xi_raw\n{lbl}", f"xi_eff\n{lbl}", f"c_adj\n{lbl}",
                    f"rank\n{lbl}", f"markup%\n{lbl}"]
        widths += [10, 10, 12, 8, 10]
    _hdr(ws, headers, widths)

    sorted_cr = sorted(countries, key=lambda c: c['cr_rank'])
    for i, c in enumerate(sorted_cr, 2):
        fill = GREEN_FILL if c['is_dev'] else None
        col = 1
        _c(ws, i, col, c['name'], fill=fill)
        col += 1
        for cfg in CONFIGS:
            lbl = cfg['label']
            _c(ws, i, col, c[f'xi_raw_{lbl}'], XI_FMT, fill=fill)
            col += 1
            _c(ws, i, col, c[f'xi_eff_{lbl}'], XI_FMT, fill=fill)
            col += 1
            _c(ws, i, col, c[f'c_adj_{lbl}'], ADJ_FMT, fill=fill)
            col += 1
            _c(ws, i, col, c[f'rank_{lbl}'], RANK_FMT, fill=fill)
            col += 1
            _c(ws, i, col, c[f'markup_{lbl}'] / 100, PCT_FMT, fill=fill)
            col += 1

    # ================================================================
    # Sheet 3: Summary (one row per config)
    # ================================================================
    ws = wb.create_sheet("Summary")
    s_headers = [
        "Config", "omega", "floor", "rho", "form",
        "Dev top5", "Dev top10", "Dev top15", "Dev top20",
    ]
    for tn in TRACKED_DEV:
        s_headers.append(f"{tn}\nrank")
    s_headers += ["Max markup%", "Min xi_eff",
                  "Spearman vs cr", "Top 5 countries"]
    s_widths = [12, 8, 8, 8, 8, 10, 10, 10, 10]
    s_widths += [12] * len(TRACKED_DEV)
    s_widths += [12, 10, 12, 40]
    _hdr(ws, s_headers, s_widths)

    cr_ranks_list = [c['cr_rank'] for c in countries]
    for r, cfg in enumerate(CONFIGS, 2):
        lbl = cfg['label']
        omega = cfg.get('omega', '')
        floor = cfg.get('floor', '')
        rho = cfg.get('rho', '')
        form = cfg['form']

        col = 1
        _c(ws, r, col, lbl, font=BOLD)
        col += 1
        _c(ws, r, col, omega if omega != '' else '')
        col += 1
        _c(ws, r, col, floor if floor != '' else '')
        col += 1
        _c(ws, r, col, rho if rho != '' else '', COST_FMT)
        col += 1
        _c(ws, r, col, form)
        col += 1

        # Dev counts
        for top_n in [5, 10, 15, 20]:
            nd = sum(1 for c in countries
                     if c[f'rank_{lbl}'] <= top_n and c['is_dev'])
            fill = BLUE_FILL if (top_n == 15 and 3 <= nd <= 8) else None
            _c(ws, r, col, nd, RANK_FMT, fill=fill)
            col += 1

        # Tracked country ranks
        for tn in TRACKED_DEV:
            tc = next((c for c in countries if c['name'] == tn), None)
            if tc:
                rank_val = tc[f'rank_{lbl}']
                fill = GREEN_FILL if rank_val <= 15 else None
                _c(ws, r, col, rank_val, RANK_FMT, fill=fill)
            else:
                _c(ws, r, col, 'N/A')
            col += 1

        # Max markup, min xi_eff
        markups = [c[f'markup_{lbl}'] for c in countries]
        xi_effs = [c[f'xi_eff_{lbl}'] for c in countries]
        _c(ws, r, col, max(markups) / 100, PCT_FMT)
        col += 1
        _c(ws, r, col, min(xi_effs), XI_FMT)
        col += 1

        # Spearman vs cr
        cfg_ranks = [c[f'rank_{lbl}'] for c in countries]
        rho_s = spearman_rho(cr_ranks_list, cfg_ranks)
        _c(ws, r, col, rho_s, RHO_FMT)
        col += 1

        # Top 5
        top5 = sorted(countries, key=lambda c: c[f'rank_{lbl}'])[:5]
        top5_str = ", ".join(
            c['name'] + (" \u2605" if c['is_dev'] else "")
            for c in top5
        )
        _c(ws, r, col, top5_str)

    # ================================================================
    # Sheet 4: Top20 (selected configs side by side)
    # ================================================================
    ws = wb.create_sheet("Top20")
    headers = []
    widths = []
    for lbl in TOP20_LABELS:
        headers += [f"Rank\n{lbl}", f"Country\n{lbl}", f"c_adj\n{lbl}"]
        widths += [8, 28, 14]
    _hdr(ws, headers, widths)

    for cfg_idx, lbl in enumerate(TOP20_LABELS):
        top20 = sorted(countries, key=lambda c: c[f'rank_{lbl}'])[:20]
        base_col = cfg_idx * 3 + 1
        for i, c in enumerate(top20, 2):
            star = " \u2605" if c['is_dev'] else ""
            _c(ws, i, base_col, c[f'rank_{lbl}'], RANK_FMT)
            _c(ws, i, base_col + 1, c['name'] + star)
            _c(ws, i, base_col + 2, c[f'c_adj_{lbl}'], ADJ_FMT)

    # ================================================================
    # Sheet 5: Stability (12 main-grid configs)
    # ================================================================
    ws = wb.create_sheet("Stability")
    headers = ["Country", "cr_rank"]
    widths = [28, 8]
    for lbl in MAIN_LABELS:
        headers.append(f"rank\n{lbl}")
        widths.append(8)
    headers += ["rank_range", "mean_rank", "top15_configs"]
    widths += [10, 10, 30]
    _hdr(ws, headers, widths)

    sorted_cr = sorted(countries, key=lambda c: c['cr_rank'])
    for i, c in enumerate(sorted_cr, 2):
        col = 1
        _c(ws, i, col, c['name'])
        col += 1
        _c(ws, i, col, c['cr_rank'], RANK_FMT)
        col += 1

        ranks = []
        top15_cfgs = []
        for lbl in MAIN_LABELS:
            r = c[f'rank_{lbl}']
            ranks.append(r)
            fill = GREEN_FILL if r <= 15 else None
            _c(ws, i, col, r, RANK_FMT, fill=fill)
            col += 1
            if r <= 15:
                top15_cfgs.append(lbl)

        rng = max(ranks) - min(ranks)
        mean_r = sum(ranks) / len(ranks)
        fill_rng = YELLOW_FILL if rng >= 20 else None
        _c(ws, i, col, rng, RANK_FMT, fill=fill_rng)
        col += 1
        _c(ws, i, col, mean_r, '0.0')
        col += 1
        _c(ws, i, col, ", ".join(top15_cfgs) if top15_cfgs else "")

    wb.save(str(OUT_XLSX))
    print(f"  Saved: {OUT_XLSX}")


# =====================================================================
# MAIN
# =====================================================================
def main():
    print("=" * 60)
    print("form_b_simulations.py  -  Form B (omega, floor) Grid")
    print("=" * 60)

    print("\n[1] Loading data...")
    countries = load_data()

    # Quick summary
    G_vals = [c['G'] for c in countries]
    R_vals = [c['R'] for c in countries]
    print(f"  G range: [{min(G_vals):.4f}, {max(G_vals):.4f}]")
    print(f"  R range: [{min(R_vals):.4f}, {max(R_vals):.4f}]")
    perfect = [c for c in countries
               if abs(c['G'] - 1.0) < 0.01 and abs(c['R'] - 1.0) < 0.01]
    print(f"  Perfect governance (G~1, R~1): "
          f"{[c['name'] for c in perfect]}")

    print("\n[2] Computing 20 configurations...")
    compute_configs(countries)

    # Print summary for key configs
    print("\n  Config summaries:")
    for lbl in ['C2', 'C3', 'A3', 'REF_A', 'REF_CR', 'REF_Bv26']:
        dev15 = sum(1 for c in countries
                    if c[f'rank_{lbl}'] <= 15 and c['is_dev'])
        top5 = sorted(countries, key=lambda c: c[f'rank_{lbl}'])[:5]
        names = [f"{c['name']}" + (" *" if c['is_dev'] else "")
                 for c in top5]
        markups = [c[f'markup_{lbl}'] for c in countries]
        print(f"    {lbl:10s}: dev15={dev15}, max_markup={max(markups):.1f}%, "
              f"top5={names}")

    print("\n[3] Validating...")
    ok = validate(countries)
    if not ok:
        print("*** VALIDATION ISSUES — see above ***\n")

    print("[4] Writing Excel...")
    write_excel(countries)

    print("\nDone.")


if __name__ == "__main__":
    main()
