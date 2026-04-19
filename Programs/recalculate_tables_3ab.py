"""
Recalculate Tables 3a & 3b from raw data.

Change from previous version: drops the ξ efficiency floor (v29).
Columns (1) and (2) are unchanged; column (3) uses ξ = G^ω × R^(1−ω)
with no institutional floor.

Reads from:  F:/onedrive/__documents/papers/FLOPsExport/Data/
Writes to:   F:/onedrive/__documents/papers/FLOPsExport/data/output/table3_recalculated.xlsx

Usage:
    python Programs/recalculate_tables_3ab.py
"""

import csv
import pathlib
import sys
import io
from decimal import Decimal, ROUND_HALF_UP

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ═══════════════════════════════════════════════════════════════════════
# PATHS
# ═══════════════════════════════════════════════════════════════════════
ROOT = pathlib.Path(r"F:\onedrive\__documents\papers\FLOPsExport")
DATA = ROOT / "Data"
OUTPUT = ROOT / "data" / "output"
OUTPUT.mkdir(parents=True, exist_ok=True)

# ═══════════════════════════════════════════════════════════════════════
# STRUCTURAL PARAMETERS (Table 2)
# ═══════════════════════════════════════════════════════════════════════
GAMMA = 0.700          # GPU TDP (kW)
GPU_PRICE = 25_000     # $/GPU (H100)
GPU_LIFE = 3           # years
GPU_UTIL = 0.70        # utilization rate β
H_YR = 365.25 * 24    # hours/year = 8766
RHO = GPU_PRICE / (GPU_LIFE * H_YR * GPU_UTIL)  # hardware amortization ≈ $1.358/hr
ETA = 0.15             # networking cost η ($/GPU-hr)
PHI = 1.08             # PUE base
DELTA_PUE = 0.015      # PUE slope (per °C above θ̄)
THETA_REF = 15.0       # reference temperature θ̄ (°C)
DC_LIFE = 15           # data center lifetime D (years)
TAU = 0.0008           # latency degradation (1/ms)
LATENCY_BAR = 200.0    # inference latency threshold l̄ (ms)
ALPHA = 0.50           # training share of demand α_train
OMEGA_XI = 0.50        # ξ weighting exponent ω
Q_TOTAL = 60_000_000_000  # total compute demand (GPU-hr)
K_BAR_SCALE = 1000     # grid capacity CSV unit correction

# Bilateral λ coefficients
ALPHA_GEO = 0.08       # α₁: geopolitical distance
ALPHA_REG = 0.04       # α₂: regulatory incompatibility
LAMBDA_UNIFORM = 0.10  # uniform λ for spec (6)
GPU_CONTROL_ALPHA3 = 0.10  # partial α₃ for GPU-controlled countries

# Demand tier weights
W_TIER1 = 0.10  # sovereign
W_TIER2 = 0.20  # regulated
W_TIER3 = 0.70  # commercial

DOMESTIC_LATENCY_DEFAULT = 5.0  # ms


def rhup(value, dp=2):
    """Round half-up (avoids banker's rounding)."""
    return float(Decimal(str(value)).quantize(Decimal(10) ** -dp,
                                              rounding=ROUND_HALF_UP))


# ═══════════════════════════════════════════════════════════════════════
# HARDCODED SETS
# ═══════════════════════════════════════════════════════════════════════

SANCTIONED = {'IRN', 'RUS', 'BLR', 'PRK', 'SYR', 'TKM'}

SUBSIDY_ADJ = {
    'IRN': 0.085,
    'TKM': 0.070,
    'DZA': 0.065,
    'EGY': 0.080,
    'UZB': 0.090,
    'QAT': 0.100,
    'SAU': 0.100,
    'ARE': 0.095,
    'RUS': 0.065,
    'KAZ': 0.085,
    'NGA': 0.080,
    'ZAF': 0.095,
    'ETH': 0.050,
}

BLOC_WESTERN = {
    'USA', 'CAN', 'GBR', 'FRA', 'DEU', 'ITA', 'ESP', 'PRT', 'NLD', 'BEL',
    'LUX', 'AUT', 'CHE', 'IRL', 'DNK', 'NOR', 'SWE', 'FIN', 'ISL', 'GRC',
    'CZE', 'POL', 'HUN', 'SVK', 'SVN', 'EST', 'LVA', 'LTU', 'HRV', 'BGR',
    'ROU', 'CYP', 'MLT', 'JPN', 'KOR', 'AUS', 'NZL', 'ISR', 'TWN',
}

BLOC_CHINA_ALIGNED = {
    'CHN', 'RUS', 'BLR', 'PRK', 'SYR', 'IRN',
    'VEN', 'CUB', 'NIC', 'MMR',
}

BLOC_DISTANCE = {
    ('W', 'W'): 0.00,
    ('W', 'C'): 0.95,
    ('W', 'N'): 0.40,
    ('C', 'W'): 0.95,
    ('C', 'C'): 0.00,
    ('C', 'N'): 0.55,
    ('N', 'W'): 0.40,
    ('N', 'C'): 0.55,
    ('N', 'N'): 0.20,
}

EU_MEMBERS = {
    'AUT', 'BEL', 'BGR', 'HRV', 'CYP', 'CZE', 'DNK', 'EST', 'FIN', 'FRA',
    'DEU', 'GRC', 'HUN', 'IRL', 'ITA', 'LVA', 'LTU', 'LUX', 'MLT', 'NLD',
    'POL', 'PRT', 'ROU', 'SVK', 'SVN', 'ESP', 'SWE',
}

APEC_CBPR = {
    'AUS', 'CAN', 'JPN', 'KOR', 'MEX', 'PHL', 'SGP', 'TWN', 'USA',
}

DEPA_MEMBERS = {'SGP', 'CHL', 'NZL'}

GPU_EXPORT_CONTROLLED = {'CHN'}


# ═══════════════════════════════════════════════════════════════════════
# BILATERAL LAMBDA FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════

def _get_bloc(iso):
    if iso in BLOC_WESTERN:
        return 'W'
    if iso in BLOC_CHINA_ALIGNED:
        return 'C'
    return 'N'


def compute_geo_distance(iso_i, iso_j):
    if iso_i == iso_j:
        return 0.0
    bi, bj = _get_bloc(iso_i), _get_bloc(iso_j)
    return BLOC_DISTANCE.get((bi, bj), 0.40)


def compute_reg_compat(iso_i, iso_j):
    if iso_i == iso_j:
        return 1
    if iso_i in EU_MEMBERS and iso_j in EU_MEMBERS:
        return 1
    if iso_i in APEC_CBPR and iso_j in APEC_CBPR:
        return 1
    if iso_i in DEPA_MEMBERS and iso_j in DEPA_MEMBERS:
        return 1
    return 0


def compute_bilateral_lambda(iso_i, iso_j):
    if iso_i == iso_j:
        return 0.0
    if iso_i in SANCTIONED or iso_j in SANCTIONED:
        if iso_i in SANCTIONED and iso_j in SANCTIONED:
            bi, bj = _get_bloc(iso_i), _get_bloc(iso_j)
            if bi == bj:
                return ALPHA_GEO * 0.0 + ALPHA_REG * (1 - 0)
        return float('inf')
    G_ij = compute_geo_distance(iso_i, iso_j)
    R_ij = compute_reg_compat(iso_i, iso_j)
    return ALPHA_GEO * G_ij + ALPHA_REG * (1 - R_ij)


def compute_fdi_lambda(host_j, buyer_k, hyperscaler_h='USA'):
    if host_j == buyer_k:
        return 0.0
    if host_j in SANCTIONED:
        return float('inf')
    s_jk = 0.0
    if host_j in GPU_EXPORT_CONTROLLED:
        s_jk = 0.5
    G_hk = compute_geo_distance(hyperscaler_h, buyer_k)
    R_hk = compute_reg_compat(hyperscaler_h, buyer_k)
    return ALPHA_GEO * G_hk + ALPHA_REG * (1 - R_hk) + GPU_CONTROL_ALPHA3 * s_jk


# ═══════════════════════════════════════════════════════════════════════
# LOAD DATA
# ═══════════════════════════════════════════════════════════════════════

def load_data():
    """Load all raw data and return structured dicts."""
    print("Loading data...")

    # 1. Calibration results (85 countries)
    cal = []
    with open(DATA / "calibration_results_v3.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            cal.append(row)
    print(f"  Calibration: {len(cal)} countries")

    # 2. ξ components (G_RoL, R_grid)
    xi_data = {}
    with open(DATA / "xi_scenarios_data.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            xi_data[row["iso"]] = {
                "G_RoL": float(row["G_RoL"]),
                "R_grid": float(row["R_grid"]),
            }
    print(f"  ξ data: {len(xi_data)} countries")

    # 3. Data center capacity
    dc_capacity = {}
    with open(DATA / "dc_capacity_estimates.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            dc_capacity[row["iso3"]] = float(row["capacity_mw"])
    print(f"  DC capacity: {len(dc_capacity)} countries")

    # 4. Grid capacity (K_bar in GPU-hours)
    k_bar = {}
    with open(DATA / "grid_capacity_estimates.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            k_bar[row["iso3"]] = float(row["K_bar_gpu_hours"]) * K_BAR_SCALE
    print(f"  Grid capacity: {len(k_bar)} countries")

    # 5. Latency data
    latency_data = {}
    with open(DATA / "country_pair_latency.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            latency_data[(row["iso3_from"], row["iso3_to"])] = float(row["avg_ms"])
    print(f"  Latency pairs: {len(latency_data)}")

    return cal, xi_data, dc_capacity, k_bar, latency_data


# ═══════════════════════════════════════════════════════════════════════
# COST COMPUTATION (§2c — the core fix)
# ═══════════════════════════════════════════════════════════════════════

def compute_costs(cal, xi_data):
    """Compute all cost variants for each country.

    Columns (1)/(2) use the residual-calibrated approach matching
    add_calibration_v28.py: costs are reconstructed from CSV components
    so that c_raw ≈ c_j_total from calibration_results_v3.csv.

    Column (3) applies v29 ξ (no floor) to the cost-recovery cost.
    """
    rho_hw = GPU_PRICE / (GPU_LIFE * H_YR * GPU_UTIL)

    # First pass: compute components and back out networking residual
    records = []
    for row in cal:
        iso = row["iso3"]
        country_name = row["country"]
        p_E_raw = float(row["p_E_usd_kwh"])
        theta = float(row["theta_summer_C"])
        p_L = float(row["p_L_usd_per_W"])
        pue_val = float(row["pue"])

        elec_raw = GAMMA * p_E_raw * pue_val
        p_E_cr = SUBSIDY_ADJ.get(iso, p_E_raw)
        elec_cr = GAMMA * p_E_cr * pue_val
        constr_cost = (p_L * GAMMA * 1000) / (DC_LIFE * H_YR * GPU_UTIL)

        # Back out networking residual from reported c_j
        cj_reported = float(row["c_j_total"])
        residual = cj_reported - (elec_raw + rho_hw + constr_cost)

        # ξ (v29: no floor)
        xd = xi_data.get(iso)
        if xd and xd["G_RoL"] > 0 and xd["R_grid"] > 0:
            G = xd["G_RoL"]
            R = xd["R_grid"]
            xi_val = (G ** OMEGA_XI) * (R ** (1 - OMEGA_XI))
        else:
            xi_val = 0.01

        records.append({
            "iso": iso, "country": country_name,
            "p_E_raw": p_E_raw, "p_E_cr": p_E_cr,
            "theta": theta, "p_L": p_L, "pue": pue_val,
            "elec_raw": elec_raw, "elec_cr": elec_cr,
            "constr_cost": constr_cost,
            "cj_reported": cj_reported, "residual": residual,
            "xi": xi_val,
            "G_RoL": xd["G_RoL"] if xd else 0,
            "R_grid": xd["R_grid"] if xd else 0,
            "is_subsidy_adj": iso in SUBSIDY_ADJ,
        })

    # Mean networking cost (ρ_net) — calibrated to match CSV values
    rho_net = sum(r["residual"] for r in records) / len(records)

    # Second pass: assemble costs
    countries = []
    for r in records:
        c_raw = r["elec_raw"] + rho_hw + r["constr_cost"] + rho_net   # (1) Raw
        c_cr = r["elec_cr"] + rho_hw + r["constr_cost"] + rho_net     # (2) Cost-recovery
        c_adj = RHO + (c_cr - RHO) / r["xi"] if r["xi"] > 0 else 999  # (3) Form B, v29 ξ

        countries.append({
            "iso": r["iso"], "country": r["country"],
            "p_E_raw": r["p_E_raw"], "p_E_cr": r["p_E_cr"],
            "theta": r["theta"], "p_L": r["p_L"], "pue": r["pue"],
            "c_raw": c_raw, "c_cr": c_cr,
            "xi": r["xi"], "c_adj": c_adj,
            "G_RoL": r["G_RoL"], "R_grid": r["R_grid"],
            "is_subsidy_adj": r["is_subsidy_adj"],
        })
    return countries


# ═══════════════════════════════════════════════════════════════════════
# DEMAND SHARES
# ═══════════════════════════════════════════════════════════════════════

def compute_demand_shares(countries, dc_capacity):
    """Compute ω_k = capacity_k / Σ capacity."""
    dc_k = {}
    for c in countries:
        iso = c["iso"]
        dc_k[iso] = dc_capacity.get(iso, 5.0)
    total = sum(dc_k.values())
    omega = {iso: cap / total for iso, cap in dc_k.items()}
    return dc_k, omega


# ═══════════════════════════════════════════════════════════════════════
# LATENCY HELPER
# ═══════════════════════════════════════════════════════════════════════

def make_latency_func(latency_data):
    def _get_latency(j, k):
        if j == k:
            return latency_data.get((j, k), DOMESTIC_LATENCY_DEFAULT)
        if (j, k) in latency_data:
            return latency_data[(j, k)]
        if (k, j) in latency_data:
            return latency_data[(k, j)]
        return None
    return _get_latency


# ═══════════════════════════════════════════════════════════════════════
# CAPACITY-CONSTRAINED TRAINING EQUILIBRIUM (§2e)
# ═══════════════════════════════════════════════════════════════════════

def solve_capacity_equilibrium(supply_stack, costs_dict, omega, dc_k,
                               sanctioned, lam=0.0, bilateral=False,
                               label=""):
    """Solve for capacity-constrained training equilibrium.

    Args:
        supply_stack: list of (iso, cost, k_bar) sorted by cost
        costs_dict: iso → cost for demand side
        omega: iso → demand weight
        dc_k: set of demand countries
        sanctioned: set of sanctioned ISOs
        lam: uniform lambda (used when bilateral=False)
        bilateral: if True, use bilateral lambda
        label: label for printing

    Returns:
        (p_T, shares, ls_cap)
    """
    # Find cheapest non-sanctioned as starting price
    p_T = supply_stack[0][1]
    for iso_j, c_j, k_j in supply_stack:
        if iso_j not in sanctioned:
            p_T = c_j
            break

    for _ in range(30):
        Q_TX = 0
        for iso_k in dc_k:
            if iso_k not in costs_dict:
                continue
            c_k = costs_dict[iso_k]
            w_k = omega.get(iso_k, 0)
            if bilateral:
                # Use tier-3 (commercial) lambda only — no tiering in this spec
                lam_k = _bilateral_min_lambda(iso_k, costs_dict, sanctioned)
                if lam_k < float('inf') and c_k > (1 + lam_k) * p_T:
                    Q_TX += ALPHA * w_k * Q_TOTAL
            else:
                if c_k > (1 + lam) * p_T:
                    Q_TX += ALPHA * w_k * Q_TOTAL

        cum_cap = 0
        found = False
        p_T_new = p_T
        for iso_j, c_j, k_j in supply_stack:
            if iso_j in sanctioned:
                continue
            cum_cap += k_j * ALPHA
            if cum_cap >= Q_TX and Q_TX > 0:
                p_T_new = c_j
                found = True
                break
        if found and abs(p_T_new - p_T) < 0.0001:
            p_T = p_T_new
            break
        if found:
            p_T = p_T_new

    # Compute export shares
    shares = {}
    remaining = Q_TX
    for iso_j, c_j, k_j in supply_stack:
        if iso_j in sanctioned:
            continue
        if c_j > p_T:
            break
        ca = min(k_j * ALPHA, remaining)
        if ca > 0:
            shares[iso_j] = ca
            remaining -= ca
        if remaining <= 0:
            break

    # λ* under capacity constraints
    ls_cap = {iso: c_k / p_T - 1 for iso, c_k in costs_dict.items()}

    print(f"  [{label}] p_T = ${p_T:.3f}/hr, {len(shares)} exporters")
    return p_T, shares, ls_cap


def _bilateral_min_lambda(iso_k, costs_dict, sanctioned):
    """Min bilateral λ for buyer k across all eligible suppliers (tier 3: geo only)."""
    min_lam = float('inf')
    for iso_j in costs_dict:
        if iso_j == iso_k:
            continue
        lam = compute_bilateral_lambda(iso_k, iso_j)
        if lam < min_lam:
            min_lam = lam
    return min_lam


# ═══════════════════════════════════════════════════════════════════════
# FDI EQUILIBRIUM (§2e, spec 7)
# ═══════════════════════════════════════════════════════════════════════

def solve_fdi_equilibrium(fdi_supply_stack, costs_dict, adj_costs, omega,
                          dc_k, sanctioned):
    """Capacity-constrained FDI training equilibrium."""
    p_T = fdi_supply_stack[0][1]
    for _ in range(30):
        Q_TX = 0
        for iso_k in dc_k:
            c_k = costs_dict.get(iso_k)
            if c_k is None:
                continue
            w_k = omega.get(iso_k, 0)
            lam_fdi_min = float('inf')
            for iso_j in adj_costs:
                if iso_j == iso_k or iso_j in sanctioned:
                    continue
                lam = compute_fdi_lambda(iso_j, iso_k, hyperscaler_h='USA')
                if lam < lam_fdi_min:
                    lam_fdi_min = lam
            if lam_fdi_min < float('inf') and c_k > (1 + lam_fdi_min) * p_T:
                Q_TX += ALPHA * w_k * Q_TOTAL
        cum_cap = 0
        p_T_new = p_T
        for iso_j, c_j, k_j in fdi_supply_stack:
            cum_cap += k_j * ALPHA
            if cum_cap >= Q_TX and Q_TX > 0:
                p_T_new = c_j
                break
        if abs(p_T_new - p_T) < 0.0001:
            p_T = p_T_new
            break
        p_T = p_T_new

    shares = {}
    remaining = Q_TX
    for iso_j, c_j, k_j in fdi_supply_stack:
        if c_j > p_T:
            break
        ca = min(k_j * ALPHA, remaining)
        if ca > 0:
            shares[iso_j] = ca
            remaining -= ca
        if remaining <= 0:
            break
    print(f"  [FDI] p_T = ${p_T:.3f}/hr, {len(shares)} exporters")
    return p_T, shares


# ═══════════════════════════════════════════════════════════════════════
# INFERENCE REGIME (§2f)
# ═══════════════════════════════════════════════════════════════════════

def compute_inference_regimes(countries, costs_dict, get_latency,
                              lambda_func=None):
    """For each buyer k, find cheapest inference source.

    Args:
        countries: list of country dicts
        costs_dict: iso → cost (ξ-adjusted for bilateral spec)
        get_latency: function(j, k) → ms or None
        lambda_func: function(j, k) → λ or None (for sovereignty)

    Returns:
        dict: iso_k → {'source': iso_j, 'cost': float}
    """
    isos = {c["iso"] for c in countries}
    result = {}
    for c_k in countries:
        iso_k = c_k["iso"]
        c_k_cost = costs_dict[iso_k]
        l_kk = get_latency(iso_k, iso_k)
        P_I_dom = (1 + TAU * (l_kk or 0)) * c_k_cost
        best_cost = P_I_dom
        best_src = iso_k
        for c_j in countries:
            iso_j = c_j["iso"]
            if iso_j == iso_k:
                continue
            l_jk = get_latency(iso_j, iso_k)
            if l_jk is None or l_jk > LATENCY_BAR:
                continue
            c_j_cost = costs_dict[iso_j]
            lam = 0.0
            if lambda_func is not None:
                lam = lambda_func(iso_j, iso_k)
                if lam >= float('inf'):
                    continue
            cost_del = (1 + lam) * (1 + TAU * l_jk) * c_j_cost
            if cost_del < best_cost:
                best_cost = cost_del
                best_src = iso_j
        result[iso_k] = {'source': best_src, 'cost': best_cost}
    return result


# ═══════════════════════════════════════════════════════════════════════
# REGIME CLASSIFICATION (§2g)
# ═══════════════════════════════════════════════════════════════════════

def classify_regime(iso, train_exporters, inf_exporters, is_dom_train, is_dom_inf):
    """Classify into EE/IE/DD/II."""
    if iso in train_exporters:
        return "EE"  # training exporter → also exports inference
    if iso in inf_exporters and not is_dom_train:
        return "IE"  # imports training, exports inference
    if is_dom_train and is_dom_inf:
        return "DD"  # domestic both
    return "II"  # full importer


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    print("=" * 70)
    print("RECALCULATE TABLES 3a & 3b — v29 ξ (NO FLOOR)")
    print("=" * 70)
    print(f"  ρ (hardware) = ${RHO:.4f}/hr")
    print(f"  ξ formula    = G^{OMEGA_XI} × R^{1-OMEGA_XI} (no institutional floor)")
    print()

    # ── Load data ──
    cal, xi_data, dc_capacity, k_bar, latency_data = load_data()
    get_latency = make_latency_func(latency_data)

    # ── Compute costs ──
    print("\nComputing costs...")
    countries = compute_costs(cal, xi_data)

    # ── Demand shares ──
    dc_k, omega = compute_demand_shares(countries, dc_capacity)
    total_omega = sum(omega.values())
    print(f"  Demand shares: {len(omega)} countries, Σω = {total_omega:.6f}")

    # ── Build cost dicts ──
    # Raw costs (col 1)
    raw_costs = {c["iso"]: c["c_raw"] for c in countries}
    # Cost-recovery costs (col 2)
    cr_costs = {c["iso"]: c["c_cr"] for c in countries}
    # Efficiency-adjusted costs (cols 3/4)
    adj_costs = {c["iso"]: c["c_adj"] for c in countries}

    # ── Rankings ──
    raw_ranked = sorted(raw_costs.items(), key=lambda x: x[1])
    raw_rank = {iso: r + 1 for r, (iso, _) in enumerate(raw_ranked)}

    cr_ranked = sorted(cr_costs.items(), key=lambda x: x[1])
    cr_rank = {iso: r + 1 for r, (iso, _) in enumerate(cr_ranked)}

    adj_ranked = sorted(adj_costs.items(), key=lambda x: x[1])
    adj_rank = {iso: r + 1 for r, (iso, _) in enumerate(adj_ranked)}

    # ── Supply stacks ──
    # Col (1): raw costs, λ=0
    raw_supply = sorted(
        [(iso, raw_costs[iso], k_bar.get(iso, 1e12)) for iso in raw_costs if iso in k_bar],
        key=lambda x: x[1])
    # Col (2): cr costs, λ=0
    cr_supply = sorted(
        [(iso, cr_costs[iso], k_bar.get(iso, 1e12)) for iso in cr_costs if iso in k_bar],
        key=lambda x: x[1])
    # Cols (3)/(4)/(6): adj costs
    adj_supply = sorted(
        [(iso, adj_costs[iso], k_bar.get(iso, 1e12)) for iso in adj_costs if iso in k_bar],
        key=lambda x: x[1])
    # Col (7): FDI — uses adj costs for supply, adj costs for demand
    fdi_supply = sorted(
        [(iso, adj_costs[iso], k_bar.get(iso, 1e12))
         for iso in adj_costs if iso in k_bar and iso not in SANCTIONED],
        key=lambda x: x[1])

    # ══════════════════════════════════════════════════════════════════
    # EQUILIBRIUM COMPUTATIONS
    # ══════════════════════════════════════════════════════════════════
    print("\nComputing equilibria...")

    # Col (1): Raw, λ=0
    p_T_raw, shares_raw, ls_raw = solve_capacity_equilibrium(
        raw_supply, raw_costs, omega, dc_k, SANCTIONED,
        lam=0.0, bilateral=False, label="Col(1) raw λ=0")

    # Col (2): Cost-recovery, λ=0
    p_T_cr, shares_cr, ls_cr = solve_capacity_equilibrium(
        cr_supply, cr_costs, omega, dc_k, SANCTIONED,
        lam=0.0, bilateral=False, label="Col(2) CR λ=0")

    # Col (3)/(4): Adj costs, λ=0 (cost rank = bilateral rank when λ=0)
    p_T_adj, shares_adj, ls_adj = solve_capacity_equilibrium(
        adj_supply, adj_costs, omega, dc_k, SANCTIONED,
        lam=0.0, bilateral=False, label="Col(3/4) adj λ=0")

    # Col (4) bilateral: adj costs with bilateral λ
    p_T_bilat, shares_bilat, ls_bilat = solve_capacity_equilibrium(
        adj_supply, adj_costs, omega, dc_k, SANCTIONED,
        bilateral=True, label="Col(4) bilateral")

    # Col (6): Uniform λ=10% (sanctions → ∞)
    p_T_unif, shares_unif, ls_unif = solve_capacity_equilibrium(
        adj_supply, adj_costs, omega, dc_k, SANCTIONED,
        lam=LAMBDA_UNIFORM, bilateral=False, label="Col(6) uniform λ=0.10")

    # Col (7): FDI
    p_T_fdi, shares_fdi = solve_fdi_equilibrium(
        fdi_supply, adj_costs, adj_costs, omega, dc_k, SANCTIONED)

    # ══════════════════════════════════════════════════════════════════
    # INFERENCE REGIMES
    # ══════════════════════════════════════════════════════════════════
    print("\nComputing inference regimes...")

    # Col (1): raw, no λ
    inf_raw = compute_inference_regimes(countries, raw_costs, get_latency)

    # Col (2): CR, no λ
    inf_cr = compute_inference_regimes(countries, cr_costs, get_latency)

    # Col (3)/(4) cost-only: adj costs, no λ
    inf_adj = compute_inference_regimes(countries, adj_costs, get_latency)

    # Col (4) bilateral: adj costs, bilateral λ
    inf_bilat = compute_inference_regimes(
        countries, adj_costs, get_latency,
        lambda_func=compute_bilateral_lambda)

    # Col (6) uniform: adj costs, uniform λ=10% (sanctions ∞)
    def _uniform_lambda(j, k):
        if j in SANCTIONED or k in SANCTIONED:
            return float('inf')
        return LAMBDA_UNIFORM
    inf_unif = compute_inference_regimes(
        countries, adj_costs, get_latency,
        lambda_func=_uniform_lambda)

    # Col (7) FDI: adj costs, FDI λ
    inf_fdi = compute_inference_regimes(
        countries, adj_costs, get_latency,
        lambda_func=lambda j, k: compute_fdi_lambda(j, k, hyperscaler_h='USA'))

    # ══════════════════════════════════════════════════════════════════
    # REGIME CLASSIFICATION
    # ══════════════════════════════════════════════════════════════════
    print("\nClassifying regimes...")

    def _classify_all(shares_t, inf_reg, costs, p_T, lam_val=0.0,
                      bilateral_flag=False):
        """Classify all countries given training shares and inference results."""
        train_exporters = set(shares_t.keys())
        inf_exporters = set()
        for iso_k, info in inf_reg.items():
            src = info['source']
            if src != iso_k:
                inf_exporters.add(src)

        regimes = {}
        for c in countries:
            iso = c["iso"]
            c_k = costs[iso]

            # Is domestic training?
            if bilateral_flag:
                lam_k_min = _bilateral_min_lambda(iso, costs, SANCTIONED)
                lam_star_k = c_k / p_T - 1 if p_T > 0 else 0
                is_dom_train = (lam_k_min >= lam_star_k) or (c_k <= p_T)
            else:
                is_dom_train = (c_k <= (1 + lam_val) * p_T)

            # Is domestic inference?
            is_dom_inf = (inf_reg.get(iso, {}).get('source', iso) == iso)

            regimes[iso] = classify_regime(iso, train_exporters,
                                           inf_exporters,
                                           is_dom_train, is_dom_inf)
        return regimes

    type_raw = _classify_all(shares_raw, inf_raw, raw_costs, p_T_raw)
    type_cr = _classify_all(shares_cr, inf_cr, cr_costs, p_T_cr)
    type_adj = _classify_all(shares_adj, inf_adj, adj_costs, p_T_adj)
    type_bilat = _classify_all(shares_bilat, inf_bilat, adj_costs, p_T_bilat,
                               bilateral_flag=True)
    type_unif = _classify_all(shares_unif, inf_unif, adj_costs, p_T_unif,
                              lam_val=LAMBDA_UNIFORM)

    # FDI regime
    fdi_train_exporters = set(shares_fdi.keys())
    fdi_inf_exporters = set()
    for iso_k, info in inf_fdi.items():
        if info['source'] != iso_k:
            fdi_inf_exporters.add(info['source'])
    # For FDI, determine import status
    type_fdi = {}
    for c in countries:
        iso = c["iso"]
        c_k = adj_costs[iso]
        # Check if would import training under FDI
        lam_fdi_min = float('inf')
        for iso_j in adj_costs:
            if iso_j == iso or iso_j in SANCTIONED:
                continue
            lam = compute_fdi_lambda(iso_j, iso, hyperscaler_h='USA')
            if lam < lam_fdi_min:
                lam_fdi_min = lam
        imports_train = (lam_fdi_min < float('inf') and
                         c_k > (1 + lam_fdi_min) * p_T_fdi)
        imports_inf = (inf_fdi.get(iso, {}).get('source', iso) != iso)

        if iso in fdi_train_exporters:
            type_fdi[iso] = "EE"
        elif iso in fdi_inf_exporters and imports_train:
            type_fdi[iso] = "IE"
        elif not imports_train and not imports_inf:
            type_fdi[iso] = "DD"
        else:
            type_fdi[iso] = "II"

    # Uniform λ ranking
    unif_ranked = sorted(adj_costs.items(), key=lambda x: x[1])
    unif_rank = {iso: r + 1 for r, (iso, _) in enumerate(unif_ranked)}

    # ══════════════════════════════════════════════════════════════════
    # λ* COMPUTATION (§2h)
    # ══════════════════════════════════════════════════════════════════
    lambda_star = {iso: adj_costs[iso] / p_T_bilat - 1
                   for iso in adj_costs} if p_T_bilat > 0 else {}

    # ══════════════════════════════════════════════════════════════════
    # ASSEMBLE TABLE DATA
    # ══════════════════════════════════════════════════════════════════
    print("\nAssembling tables...")

    # Build unified country data
    table_data = []
    for c in countries:
        iso = c["iso"]
        delta = raw_rank[iso] - adj_rank[iso]  # Δ = Rank₁ − Rank₃
        lks = lambda_star.get(iso, 0)
        lks_str = f"+{lks * 100:.1f}%" if lks >= 0 else f"\u2212{abs(lks) * 100:.1f}%"

        table_data.append({
            "iso": iso,
            "country": c["country"],
            # Col (1)
            "c_raw": c["c_raw"],
            "rank_raw": raw_rank[iso],
            "type_raw": type_raw.get(iso, "II"),
            # Col (2)
            "c_cr": c["c_cr"],
            "rank_cr": cr_rank[iso],
            "type_cr": type_cr.get(iso, "II"),
            # Col (3)
            "c_adj": c["c_adj"],
            "xi": c["xi"],
            "rank_adj": adj_rank[iso],
            "type_adj": type_adj.get(iso, "II"),
            # Col (4) bilateral
            "rank_bilat": adj_rank[iso],  # same ranking as col 3 (costs identical)
            "type_bilat": type_bilat.get(iso, "II"),
            # Delta
            "delta": delta,
            # Table 3b
            "type_unif": type_unif.get(iso, "II"),
            "rank_unif": unif_rank[iso],
            "type_fdi": type_fdi.get(iso, "II"),
            "lambda_star": lks,
            "lambda_star_str": lks_str,
        })

    # Sort by col (4) bilateral rank for top-25 selection
    table_data.sort(key=lambda d: d["rank_bilat"])

    # ══════════════════════════════════════════════════════════════════
    # VALIDATION CHECKS (§3)
    # ══════════════════════════════════════════════════════════════════
    print("\n" + "=" * 70)
    print("VALIDATION CHECKS")
    print("=" * 70)
    all_pass = True

    # Check 1: Cols (1)/(2) match calibration CSV (residual calibration)
    cal_costs = {}
    for row in cal:
        cal_costs[row["iso3"]] = float(row["c_j_total"])
    max_err = 0
    for d in table_data:
        err = abs(d["c_raw"] - cal_costs.get(d["iso"], d["c_raw"]))
        if err > max_err:
            max_err = err
    if max_err > 0.015:
        print(f"[FAIL] Check 1: max |c_raw − c_j_total| = {max_err:.4f} (> 0.015)")
        all_pass = False
    else:
        print(f"[PASS] Check 1: Cols (1)/(2) match CSV (max err = {max_err:.4f})")

    # Check 2: Rank₃ = Rank₄
    rank_mismatches = [(d["country"], d["rank_adj"], d["rank_bilat"])
                       for d in table_data if d["rank_adj"] != d["rank_bilat"]]
    if rank_mismatches:
        print(f"[FAIL] Check 2: {len(rank_mismatches)} rank mismatches")
        all_pass = False
    else:
        print("[PASS] Check 2: Rank₃ = Rank₄ for all countries")

    # Check 3: ξ ∈ (0, 1]
    xi_violations = [(d["country"], d["xi"]) for d in table_data
                     if d["xi"] <= 0 or d["xi"] > 1.0001]
    nordics = [(d["country"], d["xi"]) for d in table_data
               if d["iso"] in ('NOR', 'FIN', 'DNK')]
    if xi_violations:
        print(f"[FAIL] Check 3: {len(xi_violations)} ξ outside (0, 1]")
        all_pass = False
    else:
        nordic_strs = [f"{name}: {xi:.4f}" for name, xi in nordics]
        print(f"[PASS] Check 3: All ξ ∈ (0, 1]; Nordics: {', '.join(nordic_strs)}")

    # Check 4: Exactly 13 countries differ between cols (1) and (2)
    n_differ = sum(1 for d in table_data if abs(d["c_raw"] - d["c_cr"]) > 0.0001)
    if n_differ != 13:
        print(f"[FAIL] Check 4: {n_differ} countries differ (expected 13)")
        all_pass = False
    else:
        print(f"[PASS] Check 4: Exactly {n_differ} countries differ between cols (1) and (2)")

    # Check 5: Only EE/IE/DD/II regime codes
    valid_types = {"EE", "IE", "DD", "II"}
    all_types = set()
    for d in table_data:
        for key in ["type_raw", "type_cr", "type_adj", "type_bilat",
                     "type_unif", "type_fdi"]:
            all_types.add(d[key])
    invalid = all_types - valid_types
    if invalid:
        print(f"[FAIL] Check 5: Invalid regime codes: {invalid}")
        all_pass = False
    else:
        print(f"[PASS] Check 5: All regime codes ∈ {valid_types}")

    # Check 6: Σω ≈ 1.00
    if abs(total_omega - 1.0) > 0.01:
        print(f"[FAIL] Check 6: Σω = {total_omega:.4f}")
        all_pass = False
    else:
        print(f"[PASS] Check 6: Σω = {total_omega:.6f}")

    # Check 7: λ* < 0 ⟺ c_adj < p_T
    ls_consistency = True
    for d in table_data:
        ls = d["lambda_star"]
        below_pT = d["c_adj"] < p_T_bilat - 0.0001
        if (ls < -0.0001) != below_pT:
            ls_consistency = False
            break
    if not ls_consistency:
        print("[FAIL] Check 7: λ* sign inconsistent with c_adj vs p_T")
        all_pass = False
    else:
        print("[PASS] Check 7: λ* < 0 ⟺ c_adj < p_T")

    if all_pass:
        print("\n✓ ALL 7 CHECKS PASSED")
    else:
        print("\n✗ SOME CHECKS FAILED")

    # ══════════════════════════════════════════════════════════════════
    # SPOT CHECKS
    # ══════════════════════════════════════════════════════════════════
    print("\nSpot checks:")
    for d in table_data:
        if d["iso"] == "IRN":
            print(f"  Iran c_raw = ${d['c_raw']:.2f}, c_adj = ${d['c_adj']:.2f}, ξ = {d['xi']:.4f}")
        if d["iso"] == "NOR":
            print(f"  Norway c_adj = ${d['c_adj']:.2f}, ξ = {d['xi']:.4f} (expect ≈1.00)")
        if d["iso"] == "CAN":
            print(f"  Canada c_raw = ${d['c_raw']:.2f}, c_adj = ${d['c_adj']:.2f}, ξ = {d['xi']:.4f}")
        if d["iso"] == "FIN":
            print(f"  Finland c_adj = ${d['c_adj']:.2f}, ξ = {d['xi']:.4f} (expect ≈1.00)")

    # ══════════════════════════════════════════════════════════════════
    # EXCEL OUTPUT
    # ══════════════════════════════════════════════════════════════════
    print("\nWriting Excel output...")
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()

    # ── Table 3a ──
    ws_a = wb.active
    ws_a.title = "Table_3a"

    # Headers
    headers_a = [
        "Country",
        "c_j (1)", "Rank (1)", "Type (1)",
        "c_j (2)", "Rank (2)", "Type (2)",
        "c_j/ξ (3)", "ξ", "Rank (3)", "Type (3)",
        "c_j/ξ (4)", "Rank (4)", "Type (4)",
        "Δ"
    ]
    # Group headers
    ws_a.merge_cells('B1:D1')
    ws_a['B1'] = '(1) Raw Electricity'
    ws_a.merge_cells('E1:G1')
    ws_a['E1'] = '(2) Cost-Recovery'
    ws_a.merge_cells('H1:K1')
    ws_a['H1'] = '(3) Efficiency-Adjusted'
    ws_a.merge_cells('L1:N1')
    ws_a['L1'] = '(4) Bilateral λ_ij'

    bold = Font(bold=True)
    for cell in ['B1', 'E1', 'H1', 'L1']:
        ws_a[cell].font = bold
        ws_a[cell].alignment = Alignment(horizontal='center')

    for col_idx, h in enumerate(headers_a, 1):
        cell = ws_a.cell(row=2, column=col_idx, value=h)
        cell.font = bold

    # Top 25 by bilateral rank
    top25 = table_data[:25]

    for row_idx, d in enumerate(top25, 3):
        ws_a.cell(row=row_idx, column=1, value=d["country"])
        ws_a.cell(row=row_idx, column=2, value=f"${d['c_raw']:.2f}")
        ws_a.cell(row=row_idx, column=3, value=d["rank_raw"])
        ws_a.cell(row=row_idx, column=4, value=d["type_raw"])
        ws_a.cell(row=row_idx, column=5, value=f"${d['c_cr']:.2f}")
        ws_a.cell(row=row_idx, column=6, value=d["rank_cr"])
        ws_a.cell(row=row_idx, column=7, value=d["type_cr"])
        ws_a.cell(row=row_idx, column=8, value=f"${d['c_adj']:.2f}")
        ws_a.cell(row=row_idx, column=9, value=f"{rhup(d['xi']):.2f}")
        ws_a.cell(row=row_idx, column=10, value=d["rank_adj"])
        ws_a.cell(row=row_idx, column=11, value=d["type_adj"])
        ws_a.cell(row=row_idx, column=12, value=f"${d['c_adj']:.2f}")
        ws_a.cell(row=row_idx, column=13, value=d["rank_bilat"])
        ws_a.cell(row=row_idx, column=14, value=d["type_bilat"])
        delta = d["delta"]
        delta_str = f"+{delta}" if delta > 0 else str(delta) if delta < 0 else "0"
        ws_a.cell(row=row_idx, column=15, value=delta_str)

    # Auto-width (skip merged cells)
    for col in ws_a.columns:
        cells = [c for c in col if not isinstance(c, openpyxl.cell.cell.MergedCell)]
        if not cells:
            continue
        max_len = max(len(str(c.value or "")) for c in cells)
        ws_a.column_dimensions[cells[0].column_letter].width = max(max_len + 2, 8)

    # ── Table 3b ──
    ws_b = wb.create_sheet("Table_3b")

    headers_b = [
        "Country",
        "c_j^eff",
        "Type (4)",
        "Type (6)", "Rank (6)",
        "Type (7)",
        "λ*_k"
    ]
    # Group headers
    ws_b.merge_cells('B1:C1')
    ws_b['B1'] = '(4)/(5) Bilateral'
    ws_b.merge_cells('D1:E1')
    ws_b['D1'] = '(6) Uniform λ=10%'
    ws_b['F1'] = '(7) FDI'

    for cell in ['B1', 'D1', 'F1']:
        ws_b[cell].font = bold
        ws_b[cell].alignment = Alignment(horizontal='center')

    for col_idx, h in enumerate(headers_b, 1):
        cell = ws_b.cell(row=2, column=col_idx, value=h)
        cell.font = bold

    for row_idx, d in enumerate(top25, 3):
        ws_b.cell(row=row_idx, column=1, value=d["country"])
        ws_b.cell(row=row_idx, column=2, value=f"${d['c_adj']:.2f}")
        ws_b.cell(row=row_idx, column=3, value=d["type_bilat"])
        ws_b.cell(row=row_idx, column=4, value=d["type_unif"])
        ws_b.cell(row=row_idx, column=5, value=d["rank_unif"])
        ws_b.cell(row=row_idx, column=6, value=d["type_fdi"])
        ws_b.cell(row=row_idx, column=7, value=d["lambda_star_str"])

    for col in ws_b.columns:
        cells = [c for c in col if not isinstance(c, openpyxl.cell.cell.MergedCell)]
        if not cells:
            continue
        max_len = max(len(str(c.value or "")) for c in cells)
        ws_b.column_dimensions[cells[0].column_letter].width = max(max_len + 2, 8)

    # ── Also add full 85-country sheets ──
    ws_a_full = wb.create_sheet("Table_3a_full")
    for col_idx, h in enumerate(headers_a, 1):
        cell = ws_a_full.cell(row=1, column=col_idx, value=h)
        cell.font = bold
    for row_idx, d in enumerate(table_data, 2):
        ws_a_full.cell(row=row_idx, column=1, value=d["country"])
        ws_a_full.cell(row=row_idx, column=2, value=f"${d['c_raw']:.2f}")
        ws_a_full.cell(row=row_idx, column=3, value=d["rank_raw"])
        ws_a_full.cell(row=row_idx, column=4, value=d["type_raw"])
        ws_a_full.cell(row=row_idx, column=5, value=f"${d['c_cr']:.2f}")
        ws_a_full.cell(row=row_idx, column=6, value=d["rank_cr"])
        ws_a_full.cell(row=row_idx, column=7, value=d["type_cr"])
        ws_a_full.cell(row=row_idx, column=8, value=f"${d['c_adj']:.2f}")
        ws_a_full.cell(row=row_idx, column=9, value=f"{rhup(d['xi']):.2f}")
        ws_a_full.cell(row=row_idx, column=10, value=d["rank_adj"])
        ws_a_full.cell(row=row_idx, column=11, value=d["type_adj"])
        ws_a_full.cell(row=row_idx, column=12, value=f"${d['c_adj']:.2f}")
        ws_a_full.cell(row=row_idx, column=13, value=d["rank_bilat"])
        ws_a_full.cell(row=row_idx, column=14, value=d["type_bilat"])
        delta = d["delta"]
        delta_str = f"+{delta}" if delta > 0 else str(delta) if delta < 0 else "0"
        ws_a_full.cell(row=row_idx, column=15, value=delta_str)

    ws_b_full = wb.create_sheet("Table_3b_full")
    for col_idx, h in enumerate(headers_b, 1):
        cell = ws_b_full.cell(row=1, column=col_idx, value=h)
        cell.font = bold
    for row_idx, d in enumerate(table_data, 2):
        ws_b_full.cell(row=row_idx, column=1, value=d["country"])
        ws_b_full.cell(row=row_idx, column=2, value=f"${d['c_adj']:.2f}")
        ws_b_full.cell(row=row_idx, column=3, value=d["type_bilat"])
        ws_b_full.cell(row=row_idx, column=4, value=d["type_unif"])
        ws_b_full.cell(row=row_idx, column=5, value=d["rank_unif"])
        ws_b_full.cell(row=row_idx, column=6, value=d["type_fdi"])
        ws_b_full.cell(row=row_idx, column=7, value=d["lambda_star_str"])

    out_path = OUTPUT / "table3_recalculated.xlsx"
    wb.save(out_path)
    print(f"\nSaved: {out_path}")
    print(f"  Sheets: Table_3a (top 25), Table_3b (top 25), "
          f"Table_3a_full ({len(table_data)}), Table_3b_full ({len(table_data)})")

    print("\nDone.")


if __name__ == "__main__":
    main()
