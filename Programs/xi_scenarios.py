"""
xi_scenarios.py - Generate xi scenario comparison workbook.

Robustness exercise for FLOPs Export Paper v26.
Compares country rankings under 3 alternative governance measures
(Regulatory Quality, Investment Protection Index, FDI/GDP)
vs the current v26 baseline (Rule of Law).

Output:
  xi_scenarios.xlsx (5 sheets) + xi_scenarios_data.csv (backup)
"""

import csv
import pathlib
import warnings

import numpy as np

try:
    import wbgapi as wb
    HAS_WBGAPI = True
except ImportError:
    wb = None
    HAS_WBGAPI = False
    warnings.warn("wbgapi not installed; will fall back to cached/CSV data")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =====================================================================
# PATHS
# =====================================================================
DATA = pathlib.Path(r"F:\onedrive\__documents\papers\FLOPsExport\Data")
OUT_XLSX = DATA / "xi_scenarios.xlsx"
OUT_CSV = DATA / "xi_scenarios_data.csv"
WGI_CACHE = DATA / "wgi_cache.csv"
FDI_CACHE = DATA / "fdi_cache.csv"

# =====================================================================
# CONSTANTS (from add_calibration_v26.py)
# =====================================================================
GAMMA = 0.700
GPU_PRICE = 25_000
GPU_LIFE = 3
GPU_UTIL = 0.70
DC_LIFE = 15
H_YR = 365.25 * 24
RHO_HW = GPU_PRICE / (GPU_LIFE * H_YR * GPU_UTIL)
ETA = 0.15
OMEGA_XI = 0.85
XI_FLOOR = 0.50

SUBSIDY_ADJ = {
    'IRN': 0.085, 'TKM': 0.070, 'DZA': 0.065, 'EGY': 0.080,
    'UZB': 0.090, 'QAT': 0.100, 'SAU': 0.100, 'ARE': 0.095,
    'RUS': 0.065, 'KAZ': 0.085, 'NGA': 0.080, 'ZAF': 0.095,
    'ETH': 0.050,
}

# OECD members present in the 85-country sample (38)
# Includes Croatia (HRV, acceded Dec 2024)
OECD_ISOS = {
    'AUS', 'AUT', 'BEL', 'CAN', 'CHL', 'COL', 'CZE', 'DNK', 'EST',
    'FIN', 'FRA', 'DEU', 'GRC', 'HRV', 'HUN', 'ISL', 'IRL', 'ISR',
    'ITA', 'JPN', 'KOR', 'LVA', 'LTU', 'LUX', 'MEX', 'NLD', 'NZL',
    'NOR', 'POL', 'PRT', 'SVK', 'SVN', 'ESP', 'SWE', 'CHE', 'TUR',
    'GBR', 'USA',
}

# WGI proxy mapping for countries without direct data
WGI_PROXIES = {'XKX': 'SRB', 'GRL': 'DNK'}

# Hardcoded defaults for 4 countries missing from reliability_index.csv
RELIABILITY_DEFAULTS = {
    'CYP': {'governance': 0.85, 'grid_quality': 0.95},
    'LUX': {'governance': 0.95, 'grid_quality': 0.98},
    'MLT': {'governance': 0.85, 'grid_quality': 0.95},
    'ZAF': {'governance': 0.50, 'grid_quality': 0.65},
}

# =====================================================================
# BIT DATA (Bilateral Investment Treaties with G7 + ICSID membership)
# Sources: UNCTAD Investment Policy Hub, ICSID member database
# BIT_G7: count of in-force BITs/investment chapters with G7 members
#          (USA, CAN, GBR, FRA, DEU, ITA, JPN)
# ICSID: 1 if country has ratified the ICSID Convention, 0 otherwise
# Format: {iso3: (BIT_G7_count, ICSID_member)}
# =====================================================================
BIT_DATA = {
    # G7 members: maximum protection (bilateral coverage = symmetric)
    'USA': (7, 1), 'CAN': (7, 1), 'GBR': (7, 1),
    'FRA': (7, 1), 'DEU': (7, 1), 'ITA': (7, 1), 'JPN': (7, 1),
    # EU non-G7 with US BIT (via bilateral BIT):
    # Also covered by CETA (CAN), EU-Japan EPA (JPN), EU-UK TCA (GBR),
    # and intra-EU framework (FRA, DEU, ITA)
    'BGR': (7, 1), 'HRV': (7, 1), 'CZE': (7, 1), 'EST': (7, 1),
    'LVA': (7, 1), 'LTU': (7, 1), 'POL': (7, 1), 'ROU': (7, 1),
    'SVK': (7, 1),
    # EU non-G7 without US BIT: 6 G7 partners
    # (FRA+DEU+ITA via EU, CAN via CETA, JPN via EPA, GBR via TCA)
    'AUT': (6, 1), 'BEL': (6, 1), 'CYP': (6, 1), 'DNK': (6, 1),
    'FIN': (6, 1), 'GRC': (6, 1), 'HUN': (6, 1), 'IRL': (6, 1),
    'LUX': (6, 1), 'MLT': (6, 1), 'NLD': (6, 1), 'PRT': (6, 1),
    'SVN': (6, 1), 'ESP': (6, 1), 'SWE': (6, 1),
    # Non-EU OECD
    'AUS': (5, 1),   # AUSFTA (USA), EPA (JPN), BITs with CAN, GBR, DEU
    'CHL': (6, 1),   # FTAs with USA, CAN, JPN; BITs with GBR, FRA, DEU
    'COL': (5, 1),   # TPA (USA), FTA (CAN), BITs with FRA, GBR, JPN
    'ISL': (4, 1),   # EFTA agreements; BITs with DEU, GBR, FRA, CAN
    'ISR': (5, 1),   # FTA (USA), BITs with GBR, FRA, DEU, ITA
    'KOR': (7, 1),   # KORUS (USA), BITs/FTAs with all G7
    'MEX': (7, 1),   # USMCA (USA/CAN), BITs with GBR, FRA, DEU, ITA, JPN
    'NZL': (4, 1),   # FTAs with some G7; BITs with GBR, DEU, CAN, JPN
    'NOR': (5, 1),   # EFTA; BITs with GBR, FRA, DEU, CAN, JPN
    'CHE': (6, 1),   # EFTA + bilateral BITs with most G7
    'TUR': (6, 1),   # BITs with USA, GBR, FRA, DEU, ITA, JPN
    # Non-OECD with good investment protection
    'SGP': (6, 1),   # USSFTA, BITs with GBR, FRA, DEU, JPN, CAN
    'ARE': (5, 1),   # BITs with GBR, FRA, DEU, ITA, JPN
    'SAU': (5, 1),   # BITs with FRA, DEU, ITA, JPN, GBR
    'QAT': (4, 1),   # BITs with FRA, GBR, ITA, JPN
    'MYS': (5, 1),   # BITs with GBR, DEU, FRA, JPN, ITA
    'CHN': (6, 1),   # BITs with FRA, DEU, GBR, ITA, JPN, CAN (no US BIT)
    'IDN': (5, 1),   # BITs with GBR, FRA, DEU, ITA, JPN
    'VNM': (5, 1),   # BITs with GBR, FRA, DEU, ITA, JPN
    'THA': (4, 1),   # BITs with DEU, GBR, FRA, JPN
    'PHL': (5, 1),   # BITs with GBR, FRA, DEU, ITA, JPN
    'PAK': (5, 1),   # BITs with GBR, FRA, DEU, ITA, JPN
    # Eastern Europe / Central Asia
    'UKR': (6, 1),   # BITs with USA, GBR, FRA, DEU, ITA, CAN
    'KAZ': (6, 1),   # BITs with USA, GBR, FRA, DEU, ITA, CAN
    'GEO': (4, 1),   # BIT with USA; BITs with FRA, DEU, GBR
    'ARM': (4, 1),   # BIT with USA; BITs with FRA, DEU, ITA
    'AZE': (4, 1),   # BIT with USA; BITs with FRA, GBR, ITA
    'MDA': (4, 1),   # BIT with USA; BITs with FRA, DEU, GBR
    'KGZ': (4, 1),   # BIT with USA; BITs with FRA, DEU, GBR
    'TJK': (3, 1),   # BITs with FRA, DEU, GBR
    'UZB': (4, 1),   # BITs with FRA, DEU, GBR, JPN
    'TKM': (2, 1),   # BITs with GBR, FRA
    # Balkans
    'XKX': (3, 1),   # Kosovo: newer state, BITs with DEU, FRA, GBR
    'MNE': (4, 1),   # Inherited + new BITs with FRA, DEU, GBR, ITA
    'SRB': (5, 1),   # BITs with FRA, DEU, GBR, ITA, JPN
    'BIH': (4, 1),   # BITs with FRA, DEU, GBR, ITA
    'MKD': (4, 1),   # BITs with FRA, DEU, GBR, ITA
    'ALB': (5, 1),   # BIT with USA; BITs with FRA, DEU, GBR, ITA
    # Africa
    'MAR': (6, 1),   # BIT with USA; BITs with FRA, DEU, GBR, ITA, JPN
    'SEN': (5, 1),   # BIT with USA; BITs with FRA, DEU, GBR, ITA
    'KEN': (5, 1),   # BITs with GBR, FRA, DEU, ITA, JPN
    'GHA': (4, 1),   # BITs with GBR, DEU, FRA, ITA
    'NGA': (4, 1),   # BITs with GBR, FRA, DEU, ITA
    'EGY': (6, 1),   # BIT with USA; BITs with GBR, FRA, DEU, ITA, JPN
    'ETH': (4, 1),   # BITs with GBR, FRA, DEU, ITA
    'ZAF': (3, 1),   # Terminated many BITs (Protection of Investment Act 2015)
    # Latin America
    'ARG': (6, 1),   # BIT with USA; BITs with FRA, GBR, DEU, ITA, CAN
    'BRA': (0, 0),   # No BITs with G7 in force; not ICSID member
    'DZA': (4, 1),   # BITs with FRA, ITA, DEU, GBR
    # Sanctioned / isolated
    'IRN': (1, 0),   # BIT with JPN (1970s); not ICSID member
    'RUS': (5, 0),   # BITs with GBR, FRA, DEU, ITA, JPN; never ratified ICSID
    'BLR': (3, 1),   # BITs with DEU, GBR, FRA
    # Special territories
    'GRL': (6, 1),   # Greenland: uses Denmark values
    # India: terminated most BITs in 2017; never joined ICSID
    'IND': (3, 0),
}


# =====================================================================
# HELPER FUNCTIONS
# =====================================================================
def spearman_rho(ranks_a, ranks_b):
    """Compute Spearman rank correlation coefficient."""
    n = len(ranks_a)
    if n < 2:
        return 0.0
    d_sq = sum((a - b) ** 2 for a, b in zip(ranks_a, ranks_b))
    return 1 - 6 * d_sq / (n * (n ** 2 - 1))


def assign_ranks(data, cost_key, rank_key):
    """Sort data by cost_key ascending and assign integer ranks."""
    sorted_data = sorted(data, key=lambda x: x[cost_key])
    for rank, d in enumerate(sorted_data, 1):
        d[rank_key] = rank


# =====================================================================
# STEP 1-2: LOAD BASELINE DATA
# =====================================================================
def load_baseline_data():
    """
    Load calibration data and reliability index.
    Compute cost-recovery costs and v26 xi values.
    Returns list of 85 country dicts.
    """
    # Read calibration_results_v3.csv
    cal = []
    with open(DATA / "calibration_results_v3.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            cal.append(row)
    print(f"  Loaded {len(cal)} countries from calibration_results_v3.csv")

    # Read reliability_index.csv
    rel = {}
    with open(DATA / "reliability_index.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rel[row["iso3"]] = {
                "governance": float(row["governance"]),
                "grid_quality": float(row["grid_quality"]),
            }

    # Add hardcoded defaults for missing countries
    for iso, defaults in RELIABILITY_DEFAULTS.items():
        if iso not in rel:
            rel[iso] = defaults
            print(f"  Using hardcoded defaults for {iso}: "
                  f"gov={defaults['governance']}, grid={defaults['grid_quality']}")

    # Compute cost-recovery costs (replicating v26 lines 6993-7029)
    data = []
    for r in cal:
        iso = r["iso3"]
        country = r["country"]
        p_E_raw = float(r["p_E_usd_kwh"])
        pue = float(r["pue"])
        p_L = float(r["p_L_usd_per_W"])
        cj_reported = float(r["c_j_total"])

        cr_price = SUBSIDY_ADJ.get(iso, p_E_raw)
        elec_raw = GAMMA * p_E_raw * pue
        elec_cr = GAMMA * cr_price * pue
        constr_cost = (p_L * GAMMA * 1000) / (DC_LIFE * H_YR * GPU_UTIL)

        residual = cj_reported - (elec_raw + RHO_HW + constr_cost)

        # xi from reliability index
        ri = rel.get(iso)
        if ri and ri["governance"] > 0 and ri["grid_quality"] > 0:
            gov = ri["governance"]
            grid = ri["grid_quality"]
            xi_raw = (gov ** OMEGA_XI) * (grid ** (1 - OMEGA_XI))
            xi_eff = XI_FLOOR + (1 - XI_FLOOR) * xi_raw
        else:
            gov = 0.0
            grid = 0.0
            xi_raw = 0.01
            xi_eff = XI_FLOOR + (1 - XI_FLOOR) * 0.01

        data.append({
            "iso": iso,
            "country": country,
            "p_E_raw": p_E_raw,
            "cr_price": cr_price,
            "pue": pue,
            "p_L": p_L,
            "cj_reported": cj_reported,
            "elec_raw": elec_raw,
            "elec_cr": elec_cr,
            "constr_cost": constr_cost,
            "residual": residual,
            "gov_csv": gov,
            "grid_csv": grid,
            "xi_raw_v26": xi_raw,
            "xi_eff_v26": xi_eff,
        })

    # Mean networking residual
    rho_net = sum(d["residual"] for d in data) / len(data)
    print(f"  rho_net (mean residual) = {rho_net:.6f}")

    # Compute costs
    for d in data:
        d["cj_cr"] = d["elec_cr"] + RHO_HW + d["constr_cost"] + rho_net
        d["cj_eff_v26"] = d["cj_cr"] / d["xi_eff_v26"] if d["xi_eff_v26"] > 0 else 999

    # Validate: recomputed raw cost should match reported
    for d in data:
        cj_raw = d["elec_raw"] + RHO_HW + d["constr_cost"] + rho_net
        assert abs(cj_raw - d["cj_reported"]) < 0.015, \
            f'{d["iso"]}: raw={cj_raw:.4f} vs reported={d["cj_reported"]:.4f}'

    # Assign ranks
    assign_ranks(data, "cj_cr", "rank_cr")
    assign_ranks(data, "cj_eff_v26", "rank_v26")

    return data


# =====================================================================
# STEP 3A: FETCH WGI DATA
# =====================================================================
def fetch_wgi_data(iso_list):
    """
    Fetch WGI Rule of Law and Regulatory Quality percentile ranks.
    Returns dict: {iso: {"G_RoL": float, "G_RegQ": float}}
    """
    result = {iso: {"G_RoL": None, "G_RegQ": None} for iso in iso_list}

    # Try loading from cache first
    if WGI_CACHE.exists():
        print("  Loading WGI data from cache...")
        with open(WGI_CACHE, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                iso = row["iso3"]
                if iso in result:
                    g_rol = row.get("G_RoL", "")
                    g_regq = row.get("G_RegQ", "")
                    if g_rol:
                        result[iso]["G_RoL"] = float(g_rol)
                    if g_regq:
                        result[iso]["G_RegQ"] = float(g_regq)
        cached_count = sum(1 for v in result.values()
                           if v["G_RoL"] is not None)
        if cached_count >= 80:
            print(f"  Cache has {cached_count}/{len(iso_list)} countries")
            return result

    # Try World Bank API
    if HAS_WBGAPI:
        print("  Fetching WGI data from World Bank API...")
        # Build list of codes the API recognizes
        api_codes = []
        proxy_map = {}  # api_code -> original iso
        for iso in iso_list:
            if iso in WGI_PROXIES:
                proxy = WGI_PROXIES[iso]
                proxy_map[proxy] = proxy_map.get(proxy, [])
                proxy_map[proxy].append(iso)
                api_codes.append(proxy)
            else:
                api_codes.append(iso)
        api_codes = list(set(api_codes))

        for indicator, key in [('RL.PER.RNK', 'G_RoL'),
                               ('RQ.PER.RNK', 'G_RegQ')]:
            try:
                df = wb.data.DataFrame(
                    indicator, economy=api_codes,
                    time='YR2023', labels=False,
                )
                # Try 2022 if 2023 is mostly NaN
                col = df.columns[0] if len(df.columns) > 0 else None
                if col is not None:
                    vals = df[col]
                    if vals.isna().sum() > len(vals) * 0.5:
                        print(f"    2023 mostly NaN for {indicator}, "
                              "trying 2022...")
                        df = wb.data.DataFrame(
                            indicator, economy=api_codes,
                            time='YR2022', labels=False,
                        )
                        col = df.columns[0] if len(df.columns) > 0 else None

                if col is not None:
                    for code in api_codes:
                        try:
                            val = df.loc[code, col]
                            if not np.isnan(val):
                                score = val / 100.0  # percentile -> [0,1]
                                # Assign to the code itself
                                if code in result:
                                    result[code][key] = score
                                # Also assign to proxy targets
                                if code in proxy_map:
                                    for orig in proxy_map[code]:
                                        result[orig][key] = score
                        except (KeyError, TypeError):
                            pass
                print(f"    {indicator}: fetched "
                      f"{sum(1 for v in result.values() if v[key] is not None)} "
                      f"values")
            except Exception as e:
                print(f"    Warning: API error for {indicator}: {e}")

        # Save cache
        print("  Saving WGI cache...")
        with open(WGI_CACHE, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["iso3", "G_RoL", "G_RegQ"])
            writer.writeheader()
            for iso in sorted(result.keys()):
                writer.writerow({
                    "iso3": iso,
                    "G_RoL": result[iso]["G_RoL"] or "",
                    "G_RegQ": result[iso]["G_RegQ"] or "",
                })

    # Fill remaining gaps from reliability_index.csv governance column
    rel = {}
    with open(DATA / "reliability_index.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rel[row["iso3"]] = float(row["governance"])
    for iso, defaults in RELIABILITY_DEFAULTS.items():
        if iso not in rel:
            rel[iso] = defaults["governance"]

    filled = 0
    for iso in iso_list:
        if result[iso]["G_RoL"] is None and iso in rel:
            result[iso]["G_RoL"] = rel[iso]
            filled += 1
        if result[iso]["G_RegQ"] is None:
            # Approximate RegQ from RoL (high correlation, ρ > 0.9)
            if result[iso]["G_RoL"] is not None:
                result[iso]["G_RegQ"] = result[iso]["G_RoL"]
            elif iso in rel:
                result[iso]["G_RegQ"] = rel[iso]
            filled += 1
    if filled > 0:
        print(f"  Filled {filled} gaps from reliability_index.csv fallback")

    return result


# =====================================================================
# STEP 3C: FETCH FDI DATA
# =====================================================================
def fetch_fdi_data(iso_list):
    """
    Fetch FDI net inflows (% of GDP) and apply logistic transformation.
    Returns dict: {iso: {"fdi_raw": float, "fdi_score": float}}
    """
    result = {iso: {"fdi_raw": None, "fdi_score": None} for iso in iso_list}

    # Try loading from cache
    if FDI_CACHE.exists():
        print("  Loading FDI data from cache...")
        with open(FDI_CACHE, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                iso = row["iso3"]
                if iso in result and row.get("fdi_raw"):
                    result[iso]["fdi_raw"] = float(row["fdi_raw"])
        cached = sum(1 for v in result.values() if v["fdi_raw"] is not None)
        if cached >= 80:
            print(f"  Cache has {cached}/{len(iso_list)} countries")
            # Compute scores from cached raw values
            _apply_fdi_logistic(result)
            return result

    # Try World Bank API
    if HAS_WBGAPI:
        print("  Fetching FDI data from World Bank API...")
        api_codes = []
        proxy_map = {}
        for iso in iso_list:
            if iso in WGI_PROXIES:
                proxy = WGI_PROXIES[iso]
                proxy_map.setdefault(proxy, []).append(iso)
                api_codes.append(proxy)
            else:
                api_codes.append(iso)
        api_codes = list(set(api_codes))

        try:
            df = wb.data.DataFrame(
                'BX.KLT.DINV.WD.GD.ZS', economy=api_codes,
                time='YR2023', labels=False,
            )
            col = df.columns[0] if len(df.columns) > 0 else None
            if col is not None and df[col].isna().sum() > len(df) * 0.5:
                print("    2023 mostly NaN, trying 2022...")
                df = wb.data.DataFrame(
                    'BX.KLT.DINV.WD.GD.ZS', economy=api_codes,
                    time='YR2022', labels=False,
                )
                col = df.columns[0] if len(df.columns) > 0 else None

            if col is not None:
                for code in api_codes:
                    try:
                        val = df.loc[code, col]
                        if not np.isnan(val):
                            if code in result:
                                result[code]["fdi_raw"] = val
                            if code in proxy_map:
                                for orig in proxy_map[code]:
                                    result[orig]["fdi_raw"] = val
                    except (KeyError, TypeError):
                        pass
            fetched = sum(1 for v in result.values()
                          if v["fdi_raw"] is not None)
            print(f"    Fetched {fetched}/{len(iso_list)} FDI values")
        except Exception as e:
            print(f"    Warning: FDI API error: {e}")

        # Save cache
        with open(FDI_CACHE, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["iso3", "fdi_raw"])
            writer.writeheader()
            for iso in sorted(result.keys()):
                writer.writerow({
                    "iso3": iso,
                    "fdi_raw": result[iso]["fdi_raw"]
                    if result[iso]["fdi_raw"] is not None else "",
                })

    # Fill gaps with median
    raw_vals = [v["fdi_raw"] for v in result.values()
                if v["fdi_raw"] is not None]
    if raw_vals:
        median_fdi = float(np.median(raw_vals))
        for iso in iso_list:
            if result[iso]["fdi_raw"] is None:
                result[iso]["fdi_raw"] = median_fdi
                print(f"    {iso}: using median FDI ({median_fdi:.2f}%)")
    else:
        # No FDI data at all: assign 2.0% (global median)
        for iso in iso_list:
            result[iso]["fdi_raw"] = 2.0

    _apply_fdi_logistic(result)
    return result


def _apply_fdi_logistic(result):
    """Apply logistic transformation to FDI raw values."""
    raw = np.array([v["fdi_raw"] for v in result.values()
                    if v["fdi_raw"] is not None])
    if len(raw) == 0:
        return
    median = float(np.median(raw))
    q75 = float(np.percentile(raw, 75))
    q25 = float(np.percentile(raw, 25))
    iqr = q75 - q25
    if iqr < 0.01:
        iqr = 1.0  # avoid division by zero
    k = 2.2 / iqr

    for v in result.values():
        if v["fdi_raw"] is not None:
            v["fdi_score"] = 1.0 / (1.0 + np.exp(-k * (v["fdi_raw"] - median)))
        else:
            v["fdi_score"] = 0.5


# =====================================================================
# STEP 4: BACK-COMPUTE GRID RELIABILITY R
# =====================================================================
def assign_grid_R(data):
    """
    Assign grid reliability R from the reliability_index.csv grid_quality
    column directly.

    Note: The plan originally called for back-computing R from fresh WGI
    Rule of Law, but the CSV governance values are on a different scale
    than WGI percentile ranks (e.g., NGA: CSV=0.45, WGI=0.20). Using
    fresh WGI causes R to cap at 1.0 for most developing countries,
    destroying the grid reliability signal. Using grid_csv directly
    preserves the original data and produces correct R values.
    """
    # Read grid_quality from CSV
    grid_vals = {}
    with open(DATA / "reliability_index.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            grid_vals[row["iso3"]] = float(row["grid_quality"])
    for iso, defaults in RELIABILITY_DEFAULTS.items():
        if iso not in grid_vals:
            grid_vals[iso] = defaults["grid_quality"]

    for d in data:
        iso = d["iso"]
        d["R_grid"] = grid_vals.get(iso, 0.50)


# =====================================================================
# STEP 5-6: COMPUTE SCENARIOS
# =====================================================================
def compute_scenarios(data, wgi_data, fdi_data):
    """Compute xi under 3 alternative governance scenarios."""
    for d in data:
        iso = d["iso"]
        R = d["R_grid"]
        cj_cr = d["cj_cr"]

        # Get governance proxies
        g_regq = wgi_data.get(iso, {}).get("G_RegQ", d["gov_csv"])
        bit_g7, icsid = BIT_DATA.get(iso, (3, 1))
        bit_score = bit_g7 / 7.0
        fdi_score = fdi_data.get(iso, {}).get("fdi_score", 0.5)
        fdi_raw = fdi_data.get(iso, {}).get("fdi_raw", 0.0)

        # Store raw inputs
        d["G_RoL"] = wgi_data.get(iso, {}).get("G_RoL", d["gov_csv"])
        d["G_RegQ"] = g_regq if g_regq is not None else d["gov_csv"]
        d["BIT_G7"] = bit_g7
        d["BIT_score"] = bit_score
        d["ICSID"] = icsid
        d["FDI_raw"] = fdi_raw
        d["FDI_score"] = fdi_score

        # Scenario A: Regulatory Quality
        G_A = max(d["G_RegQ"], 0.001)
        xi_raw_A = (G_A ** OMEGA_XI) * (R ** (1 - OMEGA_XI))
        xi_eff_A = XI_FLOOR + (1 - XI_FLOOR) * xi_raw_A
        d["xi_raw_A"] = xi_raw_A
        d["xi_eff_A"] = xi_eff_A
        d["adj_cost_A"] = cj_cr / xi_eff_A

        # Scenario B: Investment Protection Index
        G_B = 0.50 * bit_score + 0.25 * icsid + 0.25 * d["G_RegQ"]
        G_B = max(G_B, 0.001)
        xi_raw_B = (G_B ** OMEGA_XI) * (R ** (1 - OMEGA_XI))
        xi_eff_B = XI_FLOOR + (1 - XI_FLOOR) * xi_raw_B
        d["G_B"] = G_B
        d["xi_raw_B"] = xi_raw_B
        d["xi_eff_B"] = xi_eff_B
        d["adj_cost_B"] = cj_cr / xi_eff_B

        # Scenario C: FDI/GDP (revealed preference)
        G_C = max(fdi_score, 0.001)
        xi_raw_C = (G_C ** OMEGA_XI) * (R ** (1 - OMEGA_XI))
        xi_eff_C = XI_FLOOR + (1 - XI_FLOOR) * xi_raw_C
        d["xi_raw_C"] = xi_raw_C
        d["xi_eff_C"] = xi_eff_C
        d["adj_cost_C"] = cj_cr / xi_eff_C

    # Assign ranks
    assign_ranks(data, "adj_cost_A", "rank_A")
    assign_ranks(data, "adj_cost_B", "rank_B")
    assign_ranks(data, "adj_cost_C", "rank_C")

    # Compute deltas (positive = improved from cr to scenario)
    for d in data:
        d["delta_v26"] = d["rank_cr"] - d["rank_v26"]
        d["delta_A"] = d["rank_cr"] - d["rank_A"]
        d["delta_B"] = d["rank_cr"] - d["rank_B"]
        d["delta_C"] = d["rank_cr"] - d["rank_C"]
        # Rank shift vs v26 baseline (positive = improved under scenario)
        d["shift_A"] = d["rank_v26"] - d["rank_A"]
        d["shift_B"] = d["rank_v26"] - d["rank_B"]
        d["shift_C"] = d["rank_v26"] - d["rank_C"]


# =====================================================================
# STEP 7: DATA QUALITY CHECKS
# =====================================================================
def quality_checks(data):
    """Run data quality checks per protocol."""
    ok = True
    n = len(data)
    print(f"\n{'='*60}")
    print("DATA QUALITY CHECKS")
    print(f"{'='*60}")

    # Check 1: 85 countries
    if n != 85:
        print(f"  FAIL: {n} countries (expected 85)")
        ok = False
    else:
        print(f"  OK: {n} countries")

    # Check 2: xi_eff in (0.50, 1.00]
    for scenario, key in [("v26", "xi_eff_v26"), ("A", "xi_eff_A"),
                          ("B", "xi_eff_B"), ("C", "xi_eff_C")]:
        vals = [d[key] for d in data]
        lo, hi = min(vals), max(vals)
        if lo < 0.499 or hi > 1.001:
            print(f"  WARN: {scenario} xi_eff range [{lo:.4f}, {hi:.4f}]")
        else:
            print(f"  OK: {scenario} xi_eff in [{lo:.4f}, {hi:.4f}]")

    # Check 3: adj_cost >= cj_cr
    for scenario, key in [("v26", "cj_eff_v26"), ("A", "adj_cost_A"),
                          ("B", "adj_cost_B"), ("C", "adj_cost_C")]:
        violations = [d for d in data if d[key] < d["cj_cr"] - 0.001]
        if violations:
            print(f"  WARN: {scenario} has {len(violations)} countries "
                  f"with adj_cost < cj_cr")
        else:
            print(f"  OK: {scenario} all adj_cost >= cj_cr")

    # Check 4: Rankings 1-85, no gaps
    for scenario, key in [("cr", "rank_cr"), ("v26", "rank_v26"),
                          ("A", "rank_A"), ("B", "rank_B"), ("C", "rank_C")]:
        ranks = sorted(d[key] for d in data)
        if ranks != list(range(1, 86)):
            print(f"  FAIL: {scenario} ranks not 1-85")
            ok = False
        else:
            print(f"  OK: {scenario} ranks 1-85 complete")

    # Check 5: Flag countries with G ~ 0
    for d in data:
        for gkey, label in [("G_RoL", "RoL"), ("G_RegQ", "RegQ"),
                            ("FDI_score", "FDI")]:
            if d.get(gkey, 1.0) < 0.05:
                print(f"  NOTE: {d['iso']} has {label}={d[gkey]:.3f}")

    print(f"{'='*60}\n")
    return ok


# =====================================================================
# STEP 8: WRITE EXCEL
# =====================================================================
COST_FMT = '$#,##0.00'
XI_FMT = '0.00'
RANK_FMT = '0'
SPEARMAN_FMT = '0.0000'
PCT_FMT = '0.00'

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                         fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                       fill_type="solid")
BOLD = Font(bold=True)
HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3",
                          fill_type="solid")


def _write_header(ws, headers, widths=None):
    """Write header row with formatting."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"


def _fmt_cell(ws, row, col, value, fmt=None, font=None, fill=None):
    """Write and format a cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if fmt:
        cell.number_format = fmt
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    return cell


def write_excel(data):
    """Write xi_scenarios.xlsx with 5 sheets."""
    wb_out = openpyxl.Workbook()

    # ── Sheet 1: Data ──
    ws = wb_out.active
    ws.title = "Data"
    headers = ["Country", "ISO3", "cr_cost", "cr_rank", "G_RoL", "G_RegQ",
               "R_grid", "BIT_G7", "BIT_score", "ICSID", "FDI_GDP",
               "FDI_score", "xi_raw_v26", "xi_eff_v26"]
    widths = [25, 6, 10, 8, 8, 8, 8, 8, 8, 8, 10, 8, 10, 10]
    _write_header(ws, headers, widths)

    for i, d in enumerate(sorted(data, key=lambda x: x["rank_cr"]), 2):
        _fmt_cell(ws, i, 1, d["country"])
        _fmt_cell(ws, i, 2, d["iso"])
        _fmt_cell(ws, i, 3, d["cj_cr"], COST_FMT)
        _fmt_cell(ws, i, 4, d["rank_cr"], RANK_FMT)
        _fmt_cell(ws, i, 5, d["G_RoL"], XI_FMT)
        _fmt_cell(ws, i, 6, d["G_RegQ"], XI_FMT)
        _fmt_cell(ws, i, 7, d["R_grid"], XI_FMT)
        _fmt_cell(ws, i, 8, d["BIT_G7"], RANK_FMT)
        _fmt_cell(ws, i, 9, d["BIT_score"], XI_FMT)
        _fmt_cell(ws, i, 10, d["ICSID"], RANK_FMT)
        _fmt_cell(ws, i, 11, d["FDI_raw"], PCT_FMT)
        _fmt_cell(ws, i, 12, d["FDI_score"], XI_FMT)
        _fmt_cell(ws, i, 13, d["xi_raw_v26"], XI_FMT)
        _fmt_cell(ws, i, 14, d["xi_eff_v26"], XI_FMT)

    # ── Sheet 2: Scenario_A (Regulatory Quality) ──
    ws = wb_out.create_sheet("Scenario_A")
    headers = ["Country", "cr_cost", "cr_rank", "G_RegQ", "R",
               "xi_raw_A", "xi_eff_A", "adj_cost_A", "rank_A",
               "delta_A", "xi_eff_v26", "rank_v26", "rank_shift"]
    widths = [25, 10, 8, 8, 8, 8, 8, 10, 8, 8, 8, 8, 8]
    _write_header(ws, headers, widths)

    sorted_a = sorted(data, key=lambda x: x["rank_A"])
    v26_top15 = {d["iso"] for d in data if d["rank_v26"] <= 15}
    for i, d in enumerate(sorted_a, 2):
        is_top10 = d["rank_A"] <= 10
        font = BOLD if is_top10 else None
        _fmt_cell(ws, i, 1, d["country"], font=font)
        _fmt_cell(ws, i, 2, d["cj_cr"], COST_FMT, font=font)
        _fmt_cell(ws, i, 3, d["rank_cr"], RANK_FMT, font=font)
        _fmt_cell(ws, i, 4, d["G_RegQ"], XI_FMT, font=font)
        _fmt_cell(ws, i, 5, d["R_grid"], XI_FMT, font=font)
        _fmt_cell(ws, i, 6, d["xi_raw_A"], XI_FMT, font=font)
        _fmt_cell(ws, i, 7, d["xi_eff_A"], XI_FMT, font=font)
        _fmt_cell(ws, i, 8, d["adj_cost_A"], COST_FMT, font=font)
        _fmt_cell(ws, i, 9, d["rank_A"], RANK_FMT, font=font)
        _fmt_cell(ws, i, 10, d["delta_A"], RANK_FMT, font=font)
        _fmt_cell(ws, i, 11, d["xi_eff_v26"], XI_FMT, font=font)
        _fmt_cell(ws, i, 12, d["rank_v26"], RANK_FMT, font=font)
        shift_fill = RED_FILL if d["shift_A"] < -10 else None
        # Green: non-OECD in top 15 under scenario but not under v26
        if (d["rank_A"] <= 15 and d["iso"] not in OECD_ISOS
                and d["iso"] not in v26_top15):
            shift_fill = GREEN_FILL
        _fmt_cell(ws, i, 13, d["shift_A"], RANK_FMT, font=font,
                  fill=shift_fill)

    # ── Sheet 3: Scenario_B (Investment Protection Index) ──
    ws = wb_out.create_sheet("Scenario_B")
    headers = ["Country", "cr_cost", "cr_rank", "BIT_score", "ICSID",
               "G_RegQ", "G_B", "R", "xi_raw_B", "xi_eff_B",
               "adj_cost_B", "rank_B", "delta_B", "xi_eff_v26",
               "rank_v26", "rank_shift"]
    widths = [25, 10, 8, 8, 8, 8, 8, 8, 8, 8, 10, 8, 8, 8, 8, 8]
    _write_header(ws, headers, widths)

    sorted_b = sorted(data, key=lambda x: x["rank_B"])
    for i, d in enumerate(sorted_b, 2):
        is_top10 = d["rank_B"] <= 10
        font = BOLD if is_top10 else None
        _fmt_cell(ws, i, 1, d["country"], font=font)
        _fmt_cell(ws, i, 2, d["cj_cr"], COST_FMT, font=font)
        _fmt_cell(ws, i, 3, d["rank_cr"], RANK_FMT, font=font)
        _fmt_cell(ws, i, 4, d["BIT_score"], XI_FMT, font=font)
        _fmt_cell(ws, i, 5, d["ICSID"], RANK_FMT, font=font)
        _fmt_cell(ws, i, 6, d["G_RegQ"], XI_FMT, font=font)
        _fmt_cell(ws, i, 7, d["G_B"], XI_FMT, font=font)
        _fmt_cell(ws, i, 8, d["R_grid"], XI_FMT, font=font)
        _fmt_cell(ws, i, 9, d["xi_raw_B"], XI_FMT, font=font)
        _fmt_cell(ws, i, 10, d["xi_eff_B"], XI_FMT, font=font)
        _fmt_cell(ws, i, 11, d["adj_cost_B"], COST_FMT, font=font)
        _fmt_cell(ws, i, 12, d["rank_B"], RANK_FMT, font=font)
        _fmt_cell(ws, i, 13, d["delta_B"], RANK_FMT, font=font)
        _fmt_cell(ws, i, 14, d["xi_eff_v26"], XI_FMT, font=font)
        _fmt_cell(ws, i, 15, d["rank_v26"], RANK_FMT, font=font)
        shift_fill = RED_FILL if d["shift_B"] < -10 else None
        if (d["rank_B"] <= 15 and d["iso"] not in OECD_ISOS
                and d["iso"] not in v26_top15):
            shift_fill = GREEN_FILL
        _fmt_cell(ws, i, 16, d["shift_B"], RANK_FMT, font=font,
                  fill=shift_fill)

    # ── Sheet 4: Scenario_C (FDI/GDP) ──
    ws = wb_out.create_sheet("Scenario_C")
    headers = ["Country", "cr_cost", "cr_rank", "FDI_GDP", "FDI_score",
               "R", "xi_raw_C", "xi_eff_C", "adj_cost_C", "rank_C",
               "delta_C", "xi_eff_v26", "rank_v26", "rank_shift"]
    widths = [25, 10, 8, 10, 8, 8, 8, 8, 10, 8, 8, 8, 8, 8]
    _write_header(ws, headers, widths)

    sorted_c = sorted(data, key=lambda x: x["rank_C"])
    for i, d in enumerate(sorted_c, 2):
        is_top10 = d["rank_C"] <= 10
        font = BOLD if is_top10 else None
        _fmt_cell(ws, i, 1, d["country"], font=font)
        _fmt_cell(ws, i, 2, d["cj_cr"], COST_FMT, font=font)
        _fmt_cell(ws, i, 3, d["rank_cr"], RANK_FMT, font=font)
        _fmt_cell(ws, i, 4, d["FDI_raw"], PCT_FMT, font=font)
        _fmt_cell(ws, i, 5, d["FDI_score"], XI_FMT, font=font)
        _fmt_cell(ws, i, 6, d["R_grid"], XI_FMT, font=font)
        _fmt_cell(ws, i, 7, d["xi_raw_C"], XI_FMT, font=font)
        _fmt_cell(ws, i, 8, d["xi_eff_C"], XI_FMT, font=font)
        _fmt_cell(ws, i, 9, d["adj_cost_C"], COST_FMT, font=font)
        _fmt_cell(ws, i, 10, d["rank_C"], RANK_FMT, font=font)
        _fmt_cell(ws, i, 11, d["delta_C"], RANK_FMT, font=font)
        _fmt_cell(ws, i, 12, d["xi_eff_v26"], XI_FMT, font=font)
        _fmt_cell(ws, i, 13, d["rank_v26"], RANK_FMT, font=font)
        shift_fill = RED_FILL if d["shift_C"] < -10 else None
        if (d["rank_C"] <= 15 and d["iso"] not in OECD_ISOS
                and d["iso"] not in v26_top15):
            shift_fill = GREEN_FILL
        _fmt_cell(ws, i, 14, d["shift_C"], RANK_FMT, font=font,
                  fill=shift_fill)

    # ── Sheet 5: Comparison ──
    ws = wb_out.create_sheet("Comparison")
    headers = ["Country", "cr_cost", "cr_rank", "rank_v26", "rank_A",
               "rank_B", "rank_C", "delta_v26", "delta_A", "delta_B",
               "delta_C"]
    widths = [25, 10, 8, 8, 8, 8, 8, 8, 8, 8, 8]
    _write_header(ws, headers, widths)

    sorted_comp = sorted(data, key=lambda x: x["rank_cr"])
    for i, d in enumerate(sorted_comp, 2):
        _fmt_cell(ws, i, 1, d["country"])
        _fmt_cell(ws, i, 2, d["cj_cr"], COST_FMT)
        _fmt_cell(ws, i, 3, d["rank_cr"], RANK_FMT)
        _fmt_cell(ws, i, 4, d["rank_v26"], RANK_FMT)
        _fmt_cell(ws, i, 5, d["rank_A"], RANK_FMT)
        _fmt_cell(ws, i, 6, d["rank_B"], RANK_FMT)
        _fmt_cell(ws, i, 7, d["rank_C"], RANK_FMT)
        _fmt_cell(ws, i, 8, d["delta_v26"], RANK_FMT)
        _fmt_cell(ws, i, 9, d["delta_A"], RANK_FMT)
        _fmt_cell(ws, i, 10, d["delta_B"], RANK_FMT)
        _fmt_cell(ws, i, 11, d["delta_C"], RANK_FMT)

        # Conditional formatting on comparison sheet
        # Green: non-OECD in top 15 under scenario but not under v26
        for col_idx, rank_key in [(5, "rank_A"), (6, "rank_B"),
                                  (7, "rank_C")]:
            if (d[rank_key] <= 15 and d["iso"] not in OECD_ISOS
                    and d["iso"] not in v26_top15):
                ws.cell(row=i, column=col_idx).fill = GREEN_FILL

        # Red fill: rank_shift < -10 (fell >10 places from v26)
        for col_idx, shift_key in [(9, "shift_A"), (10, "shift_B"),
                                   (11, "shift_C")]:
            if shift_key in d and d[shift_key] < -10:
                ws.cell(row=i, column=col_idx).fill = RED_FILL

    # ── Summary block (starting at row 90) ──
    summary_row = 90

    def non_oecd(d):
        return d["iso"] not in OECD_ISOS

    # Compute summary statistics
    scenarios = {
        "v26 (RoL)": {"rank": "rank_v26", "xi": "xi_eff_v26"},
        "A (RegQ)": {"rank": "rank_A", "xi": "xi_eff_A"},
        "B (InvProt)": {"rank": "rank_B", "xi": "xi_eff_B"},
        "C (FDI)": {"rank": "rank_C", "xi": "xi_eff_C"},
    }

    metrics = []
    # Row 1: header
    metrics.append(["Metric"] + list(scenarios.keys()))
    # Row 2: Developing countries in top 10
    row_data = ["Developing in top 10"]
    for s in scenarios.values():
        count = sum(1 for d in data
                    if d[s["rank"]] <= 10 and non_oecd(d))
        row_data.append(count)
    metrics.append(row_data)
    # Row 3: Developing countries in top 15
    row_data = ["Developing in top 15"]
    for s in scenarios.values():
        count = sum(1 for d in data
                    if d[s["rank"]] <= 15 and non_oecd(d))
        row_data.append(count)
    metrics.append(row_data)
    # Row 4: Spearman rho vs cr_rank
    row_data = ["Spearman rho (vs cr_rank)"]
    for s in scenarios.values():
        cr_ranks = [d["rank_cr"] for d in data]
        s_ranks = [d[s["rank"]] for d in data]
        rho = spearman_rho(cr_ranks, s_ranks)
        row_data.append(rho)
    metrics.append(row_data)
    # Row 4b: Spearman rho vs v26 rank (robustness metric)
    row_data = ["Spearman rho (vs v26)"]
    for sname, s in scenarios.items():
        v26_ranks = [d["rank_v26"] for d in data]
        s_ranks = [d[s["rank"]] for d in data]
        rho = spearman_rho(v26_ranks, s_ranks)
        row_data.append(rho)
    metrics.append(row_data)
    # Row 5: xi_eff range
    row_data = ["xi_eff range"]
    for s in scenarios.values():
        vals = [d[s["xi"]] for d in data]
        row_data.append(f"[{min(vals):.2f}, {max(vals):.2f}]")
    metrics.append(row_data)
    # Row 6: Mean xi_eff (OECD)
    row_data = ["Mean xi_eff (OECD)"]
    for s in scenarios.values():
        vals = [d[s["xi"]] for d in data if d["iso"] in OECD_ISOS]
        row_data.append(np.mean(vals) if vals else 0)
    metrics.append(row_data)
    # Row 7: Mean xi_eff (non-OECD)
    row_data = ["Mean xi_eff (non-OECD)"]
    for s in scenarios.values():
        vals = [d[s["xi"]] for d in data if non_oecd(d)]
        row_data.append(np.mean(vals) if vals else 0)
    metrics.append(row_data)
    # Row 8: Top 5 countries
    row_data = ["Top 5 countries"]
    for s in scenarios.values():
        top5 = sorted(data, key=lambda x: x[s["rank"]])[:5]
        names = [d["iso"] for d in top5]
        row_data.append(", ".join(names))
    metrics.append(row_data)

    # Write summary block
    for r, row_data in enumerate(metrics):
        for c, val in enumerate(row_data):
            cell = ws.cell(row=summary_row + r, column=c + 1, value=val)
            if r == 0:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            elif isinstance(val, float):
                if "Spearman" in str(metrics[r][0]):
                    cell.number_format = SPEARMAN_FMT
                else:
                    cell.number_format = XI_FMT

    # Save
    wb_out.save(str(OUT_XLSX))
    print(f"  Saved: {OUT_XLSX}")


def write_csv(data):
    """Write backup CSV."""
    fields = ["iso", "country", "cj_cr", "rank_cr", "G_RoL", "G_RegQ",
              "R_grid", "BIT_G7", "BIT_score", "ICSID", "FDI_raw",
              "FDI_score", "xi_raw_v26", "xi_eff_v26", "rank_v26",
              "xi_eff_A", "adj_cost_A", "rank_A", "delta_A",
              "xi_eff_B", "adj_cost_B", "rank_B", "delta_B",
              "xi_eff_C", "adj_cost_C", "rank_C", "delta_C",
              "shift_A", "shift_B", "shift_C"]
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for d in sorted(data, key=lambda x: x["rank_cr"]):
            writer.writerow(d)
    print(f"  Saved: {OUT_CSV}")


# =====================================================================
# MAIN
# =====================================================================
def main():
    print("=" * 60)
    print("xi_scenarios.py — Governance Robustness Exercise")
    print("=" * 60)

    # Step 1-2: Load baseline
    print("\n[1] Loading baseline data...")
    data = load_baseline_data()

    # Verification: top 5 efficiency-adjusted
    top5_v26 = sorted(data, key=lambda x: x["rank_v26"])[:5]
    print("\n  v26 efficiency-adjusted top 5:")
    for d in top5_v26:
        print(f"    {d['rank_v26']:2d}. {d['iso']} ({d['country']}) "
              f"${d['cj_eff_v26']:.2f}  xi_eff={d['xi_eff_v26']:.4f}")

    # Verification: xi_raw range
    xi_raws = [d["xi_raw_v26"] for d in data]
    print(f"\n  xi_raw range: [{min(xi_raws):.4f}, {max(xi_raws):.4f}]")
    min_iso = min(data, key=lambda x: x["xi_raw_v26"])
    print(f"  Minimum xi_raw: {min_iso['iso']} ({min_iso['country']}) "
          f"= {min_iso['xi_raw_v26']:.4f}")

    iso_list = [d["iso"] for d in data]

    # Step 3A: WGI data
    print("\n[2] Fetching WGI governance data...")
    wgi_data = fetch_wgi_data(iso_list)

    # Step 3C: FDI data
    print("\n[3] Fetching FDI data...")
    fdi_data = fetch_fdi_data(iso_list)

    # Step 4: Assign grid reliability R
    print("\n[4] Assigning grid reliability R...")
    assign_grid_R(data)

    # Verification: R values
    oecd_Rs = [d["R_grid"] for d in data if d["iso"] in OECD_ISOS]
    print(f"  OECD mean R = {np.mean(oecd_Rs):.4f}")
    for iso in ["NGA", "ETH", "PAK"]:
        d = next((x for x in data if x["iso"] == iso), None)
        if d:
            print(f"  {iso} R = {d['R_grid']:.4f}")

    # Step 5: Merge governance data into records
    for d in data:
        iso = d["iso"]
        d["G_RoL"] = wgi_data.get(iso, {}).get("G_RoL", d["gov_csv"])
        d["G_RegQ"] = wgi_data.get(iso, {}).get("G_RegQ", d["gov_csv"])

    # Step 6: Compute scenarios
    print("\n[5] Computing alternative scenarios...")
    compute_scenarios(data, wgi_data, fdi_data)

    # Step 7: Quality checks
    quality_checks(data)

    # Scenario summaries
    print("  Scenario summaries (Spearman vs v26 = robustness metric):")
    for label, rank_key in [("v26 (RoL)", "rank_v26"),
                            ("A (RegQ)", "rank_A"),
                            ("B (InvProt)", "rank_B"),
                            ("C (FDI)", "rank_C")]:
        top5 = sorted(data, key=lambda x: x[rank_key])[:5]
        names = [f"{d['iso']}" for d in top5]
        rho_cr = spearman_rho([d["rank_cr"] for d in data],
                              [d[rank_key] for d in data])
        rho_v26 = spearman_rho([d["rank_v26"] for d in data],
                               [d[rank_key] for d in data])
        non_oecd_top10 = sum(1 for d in data
                             if d[rank_key] <= 10
                             and d["iso"] not in OECD_ISOS)
        print(f"    {label:15s}: top 5 = {', '.join(names):25s} "
              f"rho_cr={rho_cr:+.4f}  rho_v26={rho_v26:.4f}  "
              f"dev_top10={non_oecd_top10}")

    # Step 8: Write output
    print("\n[6] Writing output files...")
    write_excel(data)
    write_csv(data)

    print("\nDone.")


if __name__ == "__main__":
    main()
